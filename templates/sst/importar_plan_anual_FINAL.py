#!/usr/bin/env python3
"""
Script de importación del Plan Anual de Trabajo PESV 2026
Ejecutar en el servidor con: python3 importar_plan_anual_FINAL.py
"""

import openpyxl
import psycopg2
import sys
import os

# Configuración de conexión (usar las mismas credenciales de config.py)
DB_CONFIG = {
    'host': 'dpg-d4g6i23e5dus739l1c80-a.oregon-postgres.render.com',
    'database': 'soporte_tecnico_bujd',
    'user': 'soporte_tecnico_bujd_user',
    'password': '4O43zJ3NiE5NrvdeMYD3hxsXgIOWVonw',
    'port': '5432',
    'sslmode': 'require'
}

# Ruta del archivo Excel (ajustar si es necesario)
EXCEL_PATH = 'Plan_Anual_de_Trabajo_2026.xlsx'

def main():
    print("=" * 60)
    print("IMPORTACIÓN DEL PLAN ANUAL DE TRABAJO PESV 2026")
    print("=" * 60)
    
    # Verificar que el archivo existe
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Error: No se encuentra el archivo {EXCEL_PATH}")
        print("   Asegúrate de ejecutar este script en la misma carpeta que el Excel")
        sys.exit(1)
    
    # Conectar a BD
    print("\n🔌 Conectando a la base de datos...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        print("✅ Conexión exitosa a Render PostgreSQL")
    except Exception as e:
        print(f"❌ Error de conexión: {e}")
        sys.exit(1)
    
    cursor = conn.cursor()
    
    # Cargar Excel
    print(f"\n📂 Cargando archivo {EXCEL_PATH}...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb['CRONOGRAMA PESV']
        print(f"✅ Excel cargado ({ws.max_row} filas, {ws.max_column} columnas)")
    except Exception as e:
        print(f"❌ Error al cargar Excel: {e}")
        sys.exit(1)
    
    # Limpiar tabla
    print("\n🗑️  Limpiando datos anteriores del plan anual...")
    cursor.execute("DELETE FROM plan_anual_trabajo")
    conn.commit()
    print("✅ Tabla limpia y lista para importar")
    
    registros_insertados = 0
    registros_con_error = 0
    
    # Mapeo de columnas del Excel
    COL_ACTIVIDAD = 2      # B
    COL_EVIDENCIA = 3      # C
    COL_CICLO = 4          # D
    COL_ARTICULOS = 5      # E
    COL_NIVEL = 6          # F
    COL_RESPONSABLES = 7   # G
    COL_RECURSOS = 8       # H
    COL_INICIO_MESES = 9   # I (Enero empieza aquí)
    
    print("\n🔄 Importando actividades...")
    print("   (Esto puede tomar 1-2 minutos...)")
    
    # Procesar desde fila 13 (después de los encabezados)
    for fila_num in range(13, ws.max_row + 1):
        row = ws[fila_num]
        
        # Extraer actividad (columna B)
        actividad = row[COL_ACTIVIDAD - 1].value
        if not actividad or str(actividad).strip() == '':
            continue
        
        # Extraer datos básicos
        evidencia = row[COL_EVIDENCIA - 1].value
        ciclo_phva = row[COL_CICLO - 1].value
        articulos = row[COL_ARTICULOS - 1].value
        nivel = row[COL_NIVEL - 1].value
        responsables = row[COL_RESPONSABLES - 1].value
        recursos = row[COL_RECURSOS - 1].value
        
        # Extraer meses (96 valores: 12 meses × 4 semanas × 2 [P, E])
        meses_data = []
        col_actual = COL_INICIO_MESES - 1
        
        for mes in range(12):  # 12 meses
            for semana in range(4):  # 4 semanas por mes
                # P (Planificado)
                p_val = row[col_actual].value if col_actual < len(row) else None
                p = str(p_val).upper() in ['P', 'E', 'X', 'SI', 'SÍ', '1'] if p_val else False
                meses_data.append(p)
                col_actual += 1
                
                # E (Ejecutado)
                e_val = row[col_actual].value if col_actual < len(row) else None
                e = str(e_val).upper() in ['E', 'X', 'SI', 'SÍ', '1'] if e_val else False
                meses_data.append(e)
                col_actual += 1
        
        # Calcular estado y porcentaje
        semanas_planificadas = sum(1 for i in range(0, len(meses_data), 2) if meses_data[i])
        semanas_ejecutadas = sum(1 for i in range(1, len(meses_data), 2) if meses_data[i])
        
        if semanas_ejecutadas > 0:
            estado = 'completado' if semanas_ejecutadas >= semanas_planificadas else 'en_proceso'
        else:
            estado = 'pendiente'
        
        porcentaje = round((semanas_ejecutadas / semanas_planificadas * 100), 2) if semanas_planificadas > 0 else 0.0
        
        # Insertar en la base de datos
        try:
            cursor.execute("""
                INSERT INTO plan_anual_trabajo (
                    actividad, evidencia, ciclo_phva, articulos_decreto, nivel_pesv,
                    responsables, recursos,
                    enero_semana1_p, enero_semana1_e, enero_semana2_p, enero_semana2_e,
                    enero_semana3_p, enero_semana3_e, enero_semana4_p, enero_semana4_e,
                    febrero_semana1_p, febrero_semana1_e, febrero_semana2_p, febrero_semana2_e,
                    febrero_semana3_p, febrero_semana3_e, febrero_semana4_p, febrero_semana4_e,
                    marzo_semana1_p, marzo_semana1_e, marzo_semana2_p, marzo_semana2_e,
                    marzo_semana3_p, marzo_semana3_e, marzo_semana4_p, marzo_semana4_e,
                    abril_semana1_p, abril_semana1_e, abril_semana2_p, abril_semana2_e,
                    abril_semana3_p, abril_semana3_e, abril_semana4_p, abril_semana4_e,
                    mayo_semana1_p, mayo_semana1_e, mayo_semana2_p, mayo_semana2_e,
                    mayo_semana3_p, mayo_semana3_e, mayo_semana4_p, mayo_semana4_e,
                    junio_semana1_p, junio_semana1_e, junio_semana2_p, junio_semana2_e,
                    junio_semana3_p, junio_semana3_e, junio_semana4_p, junio_semana4_e,
                    julio_semana1_p, julio_semana1_e, julio_semana2_p, julio_semana2_e,
                    julio_semana3_p, julio_semana3_e, julio_semana4_p, julio_semana4_e,
                    agosto_semana1_p, agosto_semana1_e, agosto_semana2_p, agosto_semana2_e,
                    agosto_semana3_p, agosto_semana3_e, agosto_semana4_p, agosto_semana4_e,
                    septiembre_semana1_p, septiembre_semana1_e, septiembre_semana2_p, septiembre_semana2_e,
                    septiembre_semana3_p, septiembre_semana3_e, septiembre_semana4_p, septiembre_semana4_e,
                    octubre_semana1_p, octubre_semana1_e, octubre_semana2_p, octubre_semana2_e,
                    octubre_semana3_p, octubre_semana3_e, octubre_semana4_p, octubre_semana4_e,
                    noviembre_semana1_p, noviembre_semana1_e, noviembre_semana2_p, noviembre_semana2_e,
                    noviembre_semana3_p, noviembre_semana3_e, noviembre_semana4_p, noviembre_semana4_e,
                    diciembre_semana1_p, diciembre_semana1_e, diciembre_semana2_p, diciembre_semana2_e,
                    diciembre_semana3_p, diciembre_semana3_e, diciembre_semana4_p, diciembre_semana4_e,
                    estado, porcentaje_avance
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s
                )
            """, [
                str(actividad)[:500],  # Limitar longitud
                str(evidencia)[:500] if evidencia else None,
                str(ciclo_phva)[:20] if ciclo_phva else None,
                str(articulos)[:200] if articulos else None,
                str(nivel)[:100] if nivel else None,
                str(responsables)[:200] if responsables else None,
                str(recursos)[:200] if recursos else None,
                *meses_data,  # Los 96 valores booleanos
                estado,
                porcentaje
            ])
            
            registros_insertados += 1
            
            # Commit cada 20 registros para mejor rendimiento
            if registros_insertados % 20 == 0:
                conn.commit()
                print(f"   📊 Procesados: {registros_insertados} registros...")
                
        except Exception as e:
            registros_con_error += 1
            print(f"   ⚠️  Error en fila {fila_num}: {str(e)[:100]}")
            continue
    
    # Commit final
    conn.commit()
    
    # Verificar cuántos registros quedaron
    cursor.execute("SELECT COUNT(*) FROM plan_anual_trabajo")
    total_en_bd = cursor.fetchone()[0]
    
    cursor.close()
    conn.close()
    
    print("\n" + "=" * 60)
    print("✅ IMPORTACIÓN COMPLETADA")
    print("=" * 60)
    print(f"   ✔️  Registros insertados: {registros_insertados}")
    print(f"   ⚠️  Registros con error: {registros_con_error}")
    print(f"   📊 Total en base de datos: {total_en_bd}")
    print("\n🎉 ¡Listo! Ahora puedes ver las actividades en:")
    print("   👉 http://tu-servidor/sst/plan-anual")
    print("=" * 60)

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Importación cancelada por el usuario")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Error fatal: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
