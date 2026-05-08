from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, send_from_directory, make_response, Response
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from database import crear_conexion, crear_tablas, verificar_y_crear_categorias_sst, obtener_categorias_sst, obtener_contenido_sst, ejecutar_consulta, guardar_archivo_en_bd, insertar_contenido_con_archivo, obtener_archivo_desde_bd
from config import Config
from werkzeug.security import check_password_hash, generate_password_hash
import psycopg2
import json
from datetime import datetime
from werkzeug.utils import secure_filename
import os
import urllib.parse
import mimetypes
from io import BytesIO
import time
import logging
from functools import wraps

# ===== RESETEAR CONTRASEÑA DEL USUARIO ASESOR =====
def resetear_password_asesor():
    """Resetea la contraseña del usuario asesor"""
    try:
        from werkzeug.security import generate_password_hash
        
        # Buscar usuario 'asesor'
        resultado = ejecutar_consulta(
            "SELECT id, usuario, rol FROM usuarios WHERE usuario = %s",
            ('asesor',),
            fetch=True
        )
        
        if resultado:
            nueva_password = generate_password_hash('asesor123')
            ejecutar_consulta(
                "UPDATE usuarios SET password = %s WHERE usuario = %s",
                (nueva_password, 'asesor'),
                commit=True
            )
            print("=" * 50)
            print("✅ CONTRASEÑA DE ASESOR RESETEADA")
            print(f"👤 Usuario: {resultado[0][1]}")
            print(f"🔑 Nueva contraseña: asesor123")
            print(f"📋 Rol: {resultado[0][2]}")
            print("=" * 50)
        else:
            # Crear usuario asesor
            nueva_password = generate_password_hash('asesor123')
            ejecutar_consulta("""
                INSERT INTO usuarios (usuario, password, rol, modulo_principal, permisos) 
                VALUES (%s, %s, %s, %s, %s)
            """, ('asesor', nueva_password, 'soporte', 'soporte', '{"ver_fichas": true, "agregar_fichas": true, "editar_fichas": true}'), commit=True)
            print("=" * 50)
            print("✅ USUARIO ASESOR CREADO")
            print("👤 Usuario: asesor")
            print("🔑 Contraseña: asesor123")
            print("=" * 50)
            
    except Exception as e:
        print(f"❌ Error: {e}")

# Llama a la función al inicio
resetear_password_asesor()

# ===== RESETEAR CONTRASEÑA DE USUARIO RH =====
def resetear_password_rh():
    """Resetea la contraseña del usuario RH a 'rh123'"""
    try:
        from werkzeug.security import generate_password_hash
        
        # Verificar si existe el usuario
        resultado = ejecutar_consulta(
            "SELECT id, usuario FROM usuarios WHERE usuario = %s",
            ('rh',),
            fetch=True
        )
        
        if resultado:
            # Actualizar contraseña
            nueva_password = generate_password_hash('rh123')
            ejecutar_consulta(
                "UPDATE usuarios SET password = %s WHERE usuario = %s",
                (nueva_password, 'rh'),
                commit=True
            )
            print("=" * 50)
            print("✅ CONTRASEÑA DE USUARIO RH RESETEADA")
            print("👤 Usuario: rh")
            print("🔑 Nueva contraseña: rh123")
            print("=" * 50)
        else:
            # Crear usuario si no existe
            nueva_password = generate_password_hash('rh123')
            ejecutar_consulta("""
                INSERT INTO usuarios (usuario, password, rol, modulo_principal, permisos) 
                VALUES (%s, %s, %s, %s, %s)
            """, ('rh', nueva_password, 'rh', 'rh', '{"ver_rh": true, "gestionar_rh": true}'), commit=True)
            print("=" * 50)
            print("✅ USUARIO RH CREADO")
            print("👤 Usuario: rh")
            print("🔑 Contraseña: rh123")
            print("=" * 50)
            
    except Exception as e:
        print(f"❌ Error al resetear contraseña: {e}")

# Llama a la función al inicio
resetear_password_rh()

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ===== HELPER: PARSEAR PERMISOS DESDE BD =====
def parsear_permisos(raw):
    """Parsea permisos desde BD — acepta dict (JSONB) o str (TEXT/JSON)."""
    if not raw:
        return {}
    try:
        if isinstance(raw, dict):
            return raw
        return json.loads(raw)
    except Exception:
        return {}

# ===== DECORADOR DE REINTENTO PARA ERRORES SSL =====
def retry_on_ssl_error(max_retries=2, delay=3):
    """Decorador para reintentar operaciones con errores SSL"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    error_str = str(e).lower()
                    is_ssl_error = any(keyword in error_str for keyword in ['ssl', 'connection', 'closed'])
                    if is_ssl_error and attempt < max_retries - 1:
                        logger.warning(f"⚠️  Error SSL en {func.__name__} (intento {attempt + 1}), reintentando en {delay} segundos...")
                        time.sleep(delay * (attempt + 1))
                        continue
                    else:
                        logger.error(f"❌ Error en {func.__name__} después de {attempt + 1} intentos: {e}")
                        raise
            return None
        return wrapper
    return decorator

# ===== DECORADORES DE PERMISOS =====
def admin_required(f):
    """Decorador para rutas que solo pueden acceder administradores"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.rol != 'admin':
            flash('Acceso denegado. Se requieren permisos de administrador.', 'error')
            return redirect(url_for('dashboard_admin'))
        return f(*args, **kwargs)
    return decorated_function

def soporte_required(f):
    """Decorador para rutas que solo pueden acceder usuarios de soporte"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Por favor inicia sesión.', 'error')
            return redirect(url_for('login'))
        if current_user.rol not in ['admin', 'soporte']:
            flash('No tienes permisos para acceder al módulo de Soporte.', 'error')
            return redirect(url_for('dashboard_admin'))
        return f(*args, **kwargs)
    return decorated_function

def sst_required(f):
    """Decorador para rutas que solo pueden acceder usuarios de SST"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Por favor inicia sesión.', 'error')
            return redirect(url_for('login'))
        if current_user.rol not in ['admin', 'sst']:
            flash('No tienes permisos para acceder al módulo de SST.', 'error')
            return redirect(url_for('dashboard_admin'))
        return f(*args, **kwargs)
    return decorated_function

def rh_required(f):
    """Decorador para rutas que solo pueden acceder usuarios de RH"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Por favor inicia sesión.', 'error')
            return redirect(url_for('login'))
        if current_user.rol not in ['admin', 'rh']:
            flash('No tienes permisos para acceder al módulo de Recursos Humanos.', 'error')
            return redirect(url_for('dashboard_admin'))
        return f(*args, **kwargs)
    return decorated_function

# ===== CONFIGURACIÓN DE LA APLICACIÓN =====
app = Flask(__name__)
app.config.from_object(Config)

# ===== CONFIGURACIÓN PARA SUBIDA DE ARCHIVOS SST =====
app.config['UPLOAD_FOLDER_SST'] = 'static/uploads/sst'
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200MB máximo
app.config['ALLOWED_EXTENSIONS'] = {
    'pdf', 'doc', 'docx', 'ppt', 'pptx', 'xls', 'xlsx',
    'jpg', 'jpeg', 'png', 'gif', 'mp4', 'avi', 'mov', 'mkv', 'webm'
}

upload_path = app.config['UPLOAD_FOLDER_SST']
os.makedirs(upload_path, exist_ok=True)
print(f"📁 Directorio de uploads: {upload_path}")
print(f"📁 ¿Existe el directorio?: {os.path.exists(upload_path)}")

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def generar_nombre_seguro(filename):
    """Generar nombre seguro para archivo con timestamp"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    name, ext = os.path.splitext(secure_filename(filename))
    name = name.replace(' ', '_').replace('-', '_')
    return f"{timestamp}_{name}{ext}"

# ===== CONFIGURACIÓN FLASK-LOGIN =====
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicia sesión para acceder a esta página.'

# ===== CONTEXT PROCESSORS =====
@app.context_processor
def inject_now():
    return {'now': datetime.now()}

@app.context_processor
def inject_permissions():
    def tiene_permiso(permiso):
        if current_user.is_authenticated:
            if hasattr(current_user, 'permisos'):
                return current_user.permisos.get(permiso, False)
        return False

    def puede_acceder_modulo(modulo):
        if current_user.is_authenticated:
            if current_user.rol == 'admin':
                return True
            if modulo == 'sst' and current_user.rol in ['admin', 'sst']:
                return True
            if modulo == 'soporte' and current_user.rol in ['admin', 'soporte']:
                return True
            if modulo == 'rh' and current_user.rol in ['admin', 'rh']:
                return True
            if modulo == 'dashboard' and current_user.rol == 'admin':
                return True
        return False

    def obtener_modulo_principal():
        if current_user.is_authenticated:
            return getattr(current_user, 'modulo_principal', 'soporte')
        return 'soporte'

    def obtener_rol_display():
        if current_user.is_authenticated:
            rol = current_user.rol
            display_map = {
                'admin': 'Administrador',
                'sst': 'SST',
                'soporte': 'Soporte Técnico',
                'rh': 'Recursos Humanos'
            }
            return display_map.get(rol, rol.capitalize())
        return ''

    return dict(
        tiene_permiso=tiene_permiso,
        puede_acceder_modulo=puede_acceder_modulo,
        obtener_modulo_principal=obtener_modulo_principal,
        obtener_rol_display=obtener_rol_display
    )

# ===== FILTROS TEMPLATE =====
@app.template_filter('format_date')
def format_date_filter(date_value, format='%d/%m/%Y'):
    """Filtro para formatear fechas - MANEJA STRINGS, DATETIME Y NONE"""
    if date_value is None:
        return 'Sin fecha'
    try:
        if hasattr(date_value, 'strftime'):
            return date_value.strftime(format)
        if isinstance(date_value, str):
            date_str = date_value.split('.')[0] if '.' in date_value else date_value
            formats_to_try = [
                '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M',
                '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y'
            ]
            for fmt in formats_to_try:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.strftime(format)
                except ValueError:
                    continue
            if len(date_str) >= 10:
                return date_str[:10] + " " + date_str[11:16] if len(date_str) >= 16 else date_str[:10]
            return date_str
        return str(date_value)
    except Exception as e:
        logger.error(f"❌ Error en format_date_filter: {e}")
        return 'Fecha inválida'

@app.template_filter('safe_tags')
def safe_tags_filter(tags_value):
    if tags_value is None:
        return []
    if isinstance(tags_value, str):
        return [tag.strip() for tag in tags_value.split(',') if tag.strip()]
    elif isinstance(tags_value, (int, float)):
        return [str(tags_value)]
    else:
        return []

@app.template_filter('is_video_url')
def is_video_url_filter(url):
    if not url:
        return False
    video_extensions = ['.mp4', '.avi', '.mov', '.mkv', '.webm']
    return any(url.lower().endswith(ext) for ext in video_extensions)

@app.template_filter('file_extension')
def file_extension_filter(filename):
    if not filename:
        return ''
    return os.path.splitext(filename)[1].lower().replace('.', '')

# ===== MODELO DE USUARIO MEJORADO =====
class User(UserMixin):
    def __init__(self, id, usuario, rol, modulo_principal, permisos=None):
        self.id = id
        self.usuario = usuario
        self.rol_original = rol
        self.rol = self._normalizar_rol(rol)
        self.modulo_principal = modulo_principal if modulo_principal else 'soporte'
        self.redireccionar_sst = False
        # Primero los permisos de BD, luego el rol base rellena las claves que falten
        permisos_bd = permisos or {}
        self.permisos = self._obtener_permisos_base()
        # Solo se sobreescriben las claves que YA existen en los permisos base
        for clave, valor in permisos_bd.items():
            if clave in self.permisos:
                self.permisos[clave] = valor

    def _normalizar_rol(self, rol):
        if not rol:
            return 'soporte'
        rol_str = str(rol).strip().lower()
        if rol_str in ['admin', 'administrador', 'administradora', 'superadmin', 'super usuario']:
            return 'admin'
        elif rol_str in ['sst', 'seguridad', 'salud', 'salud y seguridad', 'seguridad y salud', 'seguridad laboral']:
            return 'sst'
        elif rol_str in ['rh', 'recursos humanos', 'rrhh', 'talento humano']:
            return 'rh'
        elif rol_str in ['soporte', 'tecnico', 'técnico', 'asistente', 'ayudante', 'operador', 'soporte técnico']:
            return 'soporte'
        else:
            logger.warning(f"Rol desconocido '{rol}', normalizando a 'soporte'")
            return 'soporte'

    def _obtener_permisos_base(self):
        if self.rol == 'admin':
            return {
                'ver_fichas': True, 'agregar_fichas': True, 'editar_fichas': True,
                'eliminar_fichas': True, 'cambiar_password': True, 'gestion_usuarios': True,
                'acceder_sst': True, 'gestionar_plan_anual': True, 'agregar_evidencias': True,
                'acceder_soporte': True, 'acceder_dashboard': True, 'administrar_sistema': True,
                'ver_rh': True, 'gestionar_rh': True,
            }
        elif self.rol == 'sst':
            return {
                'ver_fichas': False, 'agregar_fichas': False, 'editar_fichas': False,
                'eliminar_fichas': False, 'cambiar_password': True, 'gestion_usuarios': False,
                'acceder_sst': True, 'gestionar_plan_anual': True, 'agregar_evidencias': True,
                'acceder_soporte': False, 'acceder_dashboard': False, 'administrar_sistema': False,
                'ver_rh': False, 'gestionar_rh': False,
            }
        elif self.rol == 'soporte':
            return {
                'ver_fichas': True, 'agregar_fichas': True, 'editar_fichas': True,
                'eliminar_fichas': True, 'cambiar_password': True, 'gestion_usuarios': False,
                'acceder_sst': False, 'gestionar_plan_anual': False, 'agregar_evidencias': False,
                'acceder_soporte': True, 'acceder_dashboard': False, 'administrar_sistema': False,
                'ver_rh': False, 'gestionar_rh': False,
            }
        else:  # rh y otros
            return {
                'ver_fichas': True, 'agregar_fichas': True, 'editar_fichas': True,
                'eliminar_fichas': True, 'cambiar_password': True, 'gestion_usuarios': False,
                'acceder_sst': False, 'acceder_soporte': True, 'acceder_dashboard': False,
                'administrar_sistema': False, 'ver_rh': True, 'gestionar_rh': True,
                'gestionar_plan_anual': False, 'agregar_evidencias': False,
            }

    def puede(self, permiso):
        return self.permisos.get(permiso, False)

    def get_rol_display(self):
        display_map = {
            'admin': 'Administrador',
            'sst': 'SST (Salud y Seguridad)',
            'soporte': 'Soporte Técnico',
            'rh': 'Recursos Humanos'
        }
        return display_map.get(self.rol, self.rol.capitalize())

@login_manager.user_loader
def load_user(user_id):
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, usuario, password, rol, modulo_principal, permisos, redireccionar_sst
            FROM usuarios WHERE id = %s
        """, (user_id,))
        user_data = cursor.fetchone()
        cursor.close()
        conn.close()
        if user_data:
            permisos = parsear_permisos(user_data[5])
            user = User(user_data[0], user_data[1], user_data[3], user_data[4], permisos)
            user.redireccionar_sst = user_data[6] if len(user_data) > 6 else False
            return user
    except Exception as e:
        logger.error(f"Error al cargar usuario: {e}")
    return None

# ===== FUNCIÓN DE REDIRECCIÓN MEJORADA =====
def redirect_a_modulo_principal():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    logger.info(f"Redirigiendo usuario {current_user.usuario} (rol: {current_user.rol})")

    if current_user.rol == 'admin':
        return redirect(url_for('dashboard_admin'))
    elif current_user.rol == 'rh':
        return redirect(url_for('rh_dashboard'))
    elif current_user.rol == 'sst':
        return redirect(url_for('sst_dashboard'))
    else:
        return redirect(url_for('index'))

# ===== RUTAS DE AUTENTICACIÓN =====
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect_a_modulo_principal()
    if request.method == 'POST':
        usuario = request.form.get('usuario')
        password = request.form.get('password')
        try:
            conn = crear_conexion()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, usuario, password, rol, modulo_principal, permisos, redireccionar_sst
                FROM usuarios WHERE usuario = %s
            """, (usuario,))
            user_data = cursor.fetchone()
            cursor.close()
            conn.close()
            if user_data and user_data[2] and user_data[2].strip():
                if check_password_hash(user_data[2], password):
                    permisos = parsear_permisos(user_data[5])
                    user = User(user_data[0], user_data[1], user_data[3], user_data[4], permisos)
                    user.redireccionar_sst = user_data[6] if len(user_data) > 6 else False
                    login_user(user)
                    flash(f'¡Bienvenido {user.usuario}!', 'success')
                    logger.info(f"Login exitoso: {user.usuario}, rol: {user.rol}, módulo: {user.modulo_principal}")

                    if user.rol == 'admin':
                        return redirect(url_for('dashboard_admin'))
                    elif user.rol == 'rh':
                        return redirect(url_for('rh_dashboard'))
                    elif user.rol == 'sst':
                        return redirect(url_for('sst_dashboard'))
                    else:  # soporte o cualquier otro
                        return redirect(url_for('index'))
                else:
                    flash('Usuario o contraseña incorrectos', 'error')
            else:
                flash('Usuario no encontrado', 'error')
        except Exception as e:
            flash(f'Error al iniciar sesión: {str(e)}', 'error')
            logger.error(f"Error en login: {e}")
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Sesión cerrada correctamente', 'info')
    return redirect(url_for('login'))

@app.route('/cambiar_password', methods=['GET', 'POST'])
@login_required
def cambiar_password():
    if request.method == 'POST':
        password_actual = request.form['password_actual']
        nueva_password = request.form['nueva_password']
        confirmar_password = request.form['confirmar_password']
        if not password_actual or not nueva_password or not confirmar_password:
            flash('Todos los campos son obligatorios', 'error')
            return render_template('cambiar_password.html')
        if nueva_password != confirmar_password:
            flash('Las nuevas contraseñas no coinciden', 'error')
            return render_template('cambiar_password.html')
        if len(nueva_password) < 6:
            flash('La nueva contraseña debe tener al menos 6 caracteres', 'error')
            return render_template('cambiar_password.html')
        try:
            resultado = ejecutar_consulta(
                "SELECT password FROM usuarios WHERE id = %s", (current_user.id,), fetch=True)
            if resultado and resultado[0] and check_password_hash(resultado[0][0], password_actual):
                hash_nueva_password = generate_password_hash(nueva_password)
                ejecutar_consulta(
                    "UPDATE usuarios SET password = %s WHERE id = %s",
                    (hash_nueva_password, current_user.id), commit=True)
                flash('Contraseña actualizada correctamente', 'success')
                return redirect_a_modulo_principal()
            else:
                flash('La contraseña actual es incorrecta', 'error')
        except Exception as e:
            flash('Error al cambiar la contraseña', 'error')
            logger.error(f"Error en cambiar_password: {e}")
    return render_template('cambiar_password.html')

# ===== RUTAS DE SOPORTE TÉCNICO =====
@app.route('/')
@login_required
def index():
    """Página principal de soporte técnico"""
    if current_user.rol == 'admin':
        pass  # Admin puede pasar
    elif current_user.rol == 'sst':
        return redirect(url_for('sst_dashboard'))
    elif current_user.rol == 'rh':
        return redirect(url_for('rh_dashboard'))

    if not current_user.puede('acceder_soporte'):
        flash('No tienes permisos para acceder al módulo de soporte', 'error')
        return redirect_a_modulo_principal()

    fichas = []
    try:
        resultado = ejecutar_consulta(
            "SELECT * FROM fichas ORDER BY fecha_actualizacion DESC", fetch=True)
        for ficha in resultado or []:
            fichas.append({
                'id': ficha[0], 'categoria': ficha[1], 'problema': ficha[2],
                'descripcion': ficha[3], 'causas': ficha[4], 'solucion': ficha[5],
                'palabras_clave': ficha[6], 'fecha_creacion': ficha[7], 'fecha_actualizacion': ficha[8]
            })
    except Exception as e:
        flash('Error al cargar las fichas', 'error')
        logger.error(f"Error en index: {e}")
    return render_template('index.html', fichas=fichas, user=current_user)


@app.route('/agregar', methods=['GET', 'POST'])
@login_required
def agregar_ficha():
    """Agregar nueva ficha de soporte"""
    if current_user.rol not in ['admin', 'soporte']:
        flash('No tienes permisos para acceder al módulo de soporte', 'error')
        return redirect_a_modulo_principal()

    if not current_user.puede('agregar_fichas'):
        flash('No tienes permisos para realizar esta acción', 'error')
        return redirect(url_for('index'))

    if request.method == 'POST':
        categoria = request.form.get('categoria', '')
        problema = request.form.get('problema', '')
        descripcion = request.form.get('descripcion', '')
        causas = request.form.get('causas', '')
        solucion = request.form.get('solucion', '')
        palabras_clave = request.form.get('palabras_clave', '')

        campos_requeridos = {'categoria': categoria, 'problema': problema, 'causas': causas, 'solucion': solucion}
        campos_faltantes = [campo for campo, valor in campos_requeridos.items() if not valor]

        if campos_faltantes:
            flash('Por favor, complete todos los campos requeridos', 'error')
            return render_template('agregar_ficha.html')

        try:
            ejecutar_consulta(
                'INSERT INTO fichas (categoria, problema, descripcion, causas, solucion, palabras_clave) VALUES (%s, %s, %s, %s, %s, %s)',
                (categoria, problema, descripcion, causas, solucion, palabras_clave), commit=True)
            flash('Ficha agregada correctamente', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            flash(f'Error al agregar la ficha: {str(e)}', 'error')

    return render_template('agregar_ficha.html')


@app.route('/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_ficha(id):
    """Editar ficha existente"""
    if current_user.rol not in ['admin', 'soporte']:
        flash('No tienes permisos para acceder al módulo de soporte', 'error')
        return redirect_a_modulo_principal()

    if not current_user.puede('editar_fichas'):
        flash('No tienes permisos para realizar esta acción', 'error')
        return redirect(url_for('index'))

    ficha = None
    try:
        if request.method == 'POST':
            categoria = request.form['categoria']
            problema = request.form['problema']
            descripcion = request.form['descripcion']
            causas = request.form['causas']
            solucion = request.form['solucion']
            palabras_clave = request.form['palabras_clave']

            causas_items = [item.strip() for item in causas.split('\n') if item.strip()]
            causas_str = '|'.join(causas_items)

            ejecutar_consulta(
                'UPDATE fichas SET categoria=%s, problema=%s, descripcion=%s, causas=%s, solucion=%s, palabras_clave=%s WHERE id=%s',
                (categoria, problema, descripcion, causas_str, solucion, palabras_clave, id), commit=True)
            flash('Ficha actualizada correctamente', 'success')
            return redirect(url_for('index'))

        resultado = ejecutar_consulta("SELECT * FROM fichas WHERE id = %s", (id,), fetch=True)
        if resultado and resultado[0]:
            ficha_data = resultado[0]
            ficha = {
                'id': ficha_data[0], 'categoria': ficha_data[1], 'problema': ficha_data[2],
                'descripcion': ficha_data[3], 'causas': ficha_data[4], 'solucion': ficha_data[5],
                'palabras_clave': ficha_data[6], 'fecha_creacion': ficha_data[7], 'fecha_actualizacion': ficha_data[8]
            }
            if ficha and ficha['causas']:
                ficha['causas'] = ficha['causas'].replace('|', '\n')
    except Exception as e:
        flash('Error al cargar/editar la ficha', 'error')
        logger.error(f"Error en editar_ficha: {e}")

    if not ficha:
        flash('Ficha no encontrada', 'error')
        return redirect(url_for('index'))

    return render_template('editar_ficha.html', ficha=ficha)


@app.route('/eliminar/<int:id>')
@login_required
def eliminar_ficha(id):
    """Eliminar ficha"""
    if current_user.rol not in ['admin', 'soporte']:
        flash('No tienes permisos para acceder al módulo de soporte', 'error')
        return redirect_a_modulo_principal()

    if not current_user.puede('eliminar_fichas'):
        flash('No tienes permisos para realizar esta acción', 'error')
        return redirect(url_for('index'))

    try:
        ejecutar_consulta("DELETE FROM fichas WHERE id = %s", (id,), commit=True)
        flash('Ficha eliminada correctamente', 'success')
    except Exception as e:
        flash('Error al eliminar la ficha', 'error')
        logger.error(f"Error en eliminar_ficha: {e}")

    return redirect(url_for('index'))


@app.route('/buscar')
@login_required
def buscar():
    """Buscar fichas"""
    if current_user.rol not in ['admin', 'soporte']:
        flash('No tienes permisos para acceder al módulo de soporte', 'error')
        return redirect_a_modulo_principal()

    if not current_user.puede('ver_fichas'):
        flash('No tienes permisos para ver las fichas', 'error')
        return redirect(url_for('index'))

    query = request.args.get('q', '')
    categoria = request.args.get('categoria', '')
    fichas = []

    try:
        if categoria and query:
            resultado = ejecutar_consulta(
                "SELECT * FROM fichas WHERE categoria = %s AND (problema LIKE %s OR palabras_clave LIKE %s)",
                (categoria, f'%{query}%', f'%{query}%'), fetch=True)
        elif categoria:
            resultado = ejecutar_consulta("SELECT * FROM fichas WHERE categoria = %s", (categoria,), fetch=True)
        elif query:
            resultado = ejecutar_consulta(
                "SELECT * FROM fichas WHERE problema LIKE %s OR palabras_clave LIKE %s",
                (f'%{query}%', f'%{query}%'), fetch=True)
        else:
            resultado = ejecutar_consulta("SELECT * FROM fichas ORDER BY fecha_actualizacion DESC", fetch=True)

        for ficha in resultado or []:
            fichas.append({
                'id': ficha[0], 'categoria': ficha[1], 'problema': ficha[2],
                'descripcion': ficha[3], 'causas': ficha[4], 'solucion': ficha[5],
                'palabras_clave': ficha[6], 'fecha_creacion': ficha[7], 'fecha_actualizacion': ficha[8]
            })
    except Exception as e:
        flash('Error en la búsqueda', 'error')
        logger.error(f"Error en buscar: {e}")

    return render_template('buscar.html', fichas=fichas, query=query, categoria=categoria)


@app.route('/ficha/<int:id>')
@login_required
def ver_ficha(id):
    """Ver detalle de ficha"""
    if current_user.rol not in ['admin', 'soporte']:
        flash('No tienes permisos para acceder al módulo de soporte', 'error')
        return redirect_a_modulo_principal()

    if not current_user.puede('ver_fichas'):
        flash('No tienes permisos para ver las fichas', 'error')
        return redirect(url_for('index'))

    ficha = None
    try:
        resultado = ejecutar_consulta("SELECT * FROM fichas WHERE id = %s", (id,), fetch=True)
        if resultado and resultado[0]:
            ficha_data = resultado[0]
            ficha = {
                'id': ficha_data[0], 'categoria': ficha_data[1], 'problema': ficha_data[2],
                'descripcion': ficha_data[3], 'causas': ficha_data[4], 'solucion': ficha_data[5],
                'palabras_clave': ficha_data[6], 'fecha_creacion': ficha_data[7], 'fecha_actualizacion': ficha_data[8]
            }
    except Exception as e:
        flash('Error al cargar la ficha', 'error')
        logger.error(f"Error en ver_ficha: {e}")

    if not ficha:
        flash('Ficha no encontrada', 'error')
        return redirect(url_for('index'))

    return render_template('ver_ficha.html', ficha=ficha)

# ===== RUTAS DE GESTIÓN DE USUARIOS (Solo Admin) =====
@app.route('/usuarios')
@login_required
def gestion_usuarios():
    if not current_user.puede('gestion_usuarios'):
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect_a_modulo_principal()
    usuarios = []
    try:
        resultado = ejecutar_consulta("""
            SELECT id, usuario, password, rol, modulo_principal, permisos, fecha_creacion, fecha_actualizacion
            FROM usuarios ORDER BY fecha_creacion DESC
        """, fetch=True)
        for usuario in resultado or []:
            usuario_dict = {
                'id': usuario[0], 'usuario': usuario[1], 'password': usuario[2],
                'rol': usuario[3], 'modulo_principal': usuario[4] if usuario[4] else 'soporte',
                'permisos': usuario[5], 'fecha_creacion': usuario[6], 'fecha_actualizacion': usuario[7]
            }
            usuario_dict['permisos_parsed'] = parsear_permisos(usuario_dict.get('permisos'))
            usuarios.append(usuario_dict)
    except Exception as e:
        flash('Error al cargar los usuarios', 'error')
        logger.error(f"❌ Error en gestion_usuarios: {e}")
    return render_template('gestion_usuarios.html', usuarios=usuarios)

@app.route('/editar_usuario/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_usuario(id):
    if not current_user.puede('gestion_usuarios'):
        flash('No tienes permisos para realizar esta acción', 'error')
        return redirect_a_modulo_principal()
    usuario_data = None
    try:
        if request.method == 'POST':
            usuario = request.form['usuario']
            password = request.form['password']
            rol = request.form['rol']
            modulo_principal = request.form['modulo_principal']
            permisos = {
                'ver_fichas': 'ver_fichas' in request.form,
                'agregar_fichas': 'agregar_fichas' in request.form,
                'editar_fichas': 'editar_fichas' in request.form,
                'eliminar_fichas': 'eliminar_fichas' in request.form,
                'cambiar_password': True
            }
            permisos_json = json.dumps(permisos)
            if password:
                hash_password = generate_password_hash(password)
                ejecutar_consulta(
                    "UPDATE usuarios SET usuario = %s, password = %s, rol = %s, modulo_principal = %s, permisos = %s WHERE id = %s",
                    (usuario, hash_password, rol, modulo_principal, permisos_json, id), commit=True)
            else:
                ejecutar_consulta(
                    "UPDATE usuarios SET usuario = %s, rol = %s, modulo_principal = %s, permisos = %s WHERE id = %s",
                    (usuario, rol, modulo_principal, permisos_json, id), commit=True)
            flash('Usuario actualizado correctamente', 'success')
            return redirect(url_for('gestion_usuarios'))
        resultado = ejecutar_consulta("SELECT * FROM usuarios WHERE id = %s", (id,), fetch=True)
        if resultado and resultado[0]:
            usuario = resultado[0]
            usuario_data = {
                'id': usuario[0], 'usuario': usuario[1], 'password': usuario[2],
                'rol': usuario[3], 'modulo_principal': usuario[4] if usuario[4] else 'soporte',
                'permisos': usuario[5], 'fecha_creacion': usuario[6], 'fecha_actualizacion': usuario[7]
            }
            usuario_data['permisos_parsed'] = parsear_permisos(usuario_data.get('permisos'))
    except psycopg2.IntegrityError:
        flash('El usuario ya existe', 'error')
    except Exception as e:
        flash('Error al editar el usuario', 'error')
        logger.error(f"Error en editar_usuario: {e}")
    if not usuario_data:
        flash('Usuario no encontrado', 'error')
        return redirect(url_for('gestion_usuarios'))
    return render_template('editar_usuario.html', usuario=usuario_data)

@app.route('/agregar_usuario', methods=['GET', 'POST'])
@login_required
def agregar_usuario():
    if not current_user.puede('gestion_usuarios'):
        flash('No tienes permisos para realizar esta acción', 'error')
        return redirect_a_modulo_principal()
    if request.method == 'POST':
        usuario = request.form['usuario']
        password = request.form['password']
        rol = request.form['rol']
        modulo_principal = request.form['modulo_principal']
        if not usuario or not password:
            flash('Usuario y contraseña son obligatorios', 'error')
            return render_template('agregar_usuario.html')
        permisos = {
            'ver_fichas': 'ver_fichas' in request.form,
            'agregar_fichas': 'agregar_fichas' in request.form,
            'editar_fichas': 'editar_fichas' in request.form,
            'eliminar_fichas': 'eliminar_fichas' in request.form,
            'cambiar_password': True
        }
        permisos_json = json.dumps(permisos)
        hash_password = generate_password_hash(password)
        try:
            ejecutar_consulta(
                "INSERT INTO usuarios (usuario, password, rol, modulo_principal, permisos) VALUES (%s, %s, %s, %s, %s)",
                (usuario, hash_password, rol, modulo_principal, permisos_json), commit=True)
            flash('Usuario agregado correctamente', 'success')
            return redirect(url_for('gestion_usuarios'))
        except psycopg2.IntegrityError:
            flash('El usuario ya existe', 'error')
        except Exception as e:
            flash('Error al agregar el usuario', 'error')
            logger.error(f"Error en agregar_usuario: {e}")
    return render_template('agregar_usuario.html')

@app.route('/eliminar_usuario/<int:id>')
@login_required
def eliminar_usuario(id):
    if not current_user.puede('gestion_usuarios'):
        flash('No tienes permisos para realizar esta acción', 'error')
        return redirect_a_modulo_principal()
    if id == current_user.id:
        flash('No puedes eliminar tu propio usuario', 'error')
        return redirect(url_for('gestion_usuarios'))
    try:
        ejecutar_consulta("DELETE FROM usuarios WHERE id = %s", (id,), commit=True)
        flash('Usuario eliminado correctamente', 'success')
    except Exception as e:
        flash('Error al eliminar el usuario', 'error')
        logger.error(f"Error en eliminar_usuario: {e}")
    return redirect(url_for('gestion_usuarios'))

# ===== RUTAS DE INFORMACIÓN =====
@app.route('/soluciones_visuales')
@login_required
def soluciones_visuales():
    soluciones = [
        {'id': 1, 'titulo': '¿Como consultamos clientes?', 'categoria': 'Softv', 'imagenes': ['softv/softv1.png', 'softv/softv2.png', 'softv/softv3.png', 'softv/softv4.png'], 'descripcion': 'Busqueda del cliente paso a paso'},
        {'id': 2, 'titulo': '¿Como vemos las facturas del usuario?', 'categoria': 'Softv', 'imagenes': ['softv/softv5.png', 'softv/softv6.png', 'softv/softv7.png', 'softv/softv8.png'], 'descripcion': 'Consultar historial de pagos del usuario'},
        {'id': 3, 'titulo': '¿Como consultamos las ordenes de servicio de los usuarios?', 'categoria': 'Softv', 'imagenes': ['softv/softv9.png', 'softv/softv10.png', 'softv/softv11.png', 'softv/softv12.png'], 'descripcion': 'Consultar historial de ordenes de servicio del usuario'},
        {'id': 4, 'titulo': '¿Como consultamos reportes de fallas de los usuarios?', 'categoria': 'Softv', 'imagenes': ['softv/softv13.png', 'softv/softv14.png', 'softv/softv15.png', 'softv/softv16.png'], 'descripcion': 'Consultar historial de reportes de falla del usuario'},
        {'id': 5, 'titulo': '¿Como creamos un reporte de falla?', 'categoria': 'Softv', 'imagenes': ['softv/softv15.png', 'softv/softv16.png', 'softv/softv17.png', 'softv/softv19.png', 'softv/softv21.png', 'softv/softv22.png'], 'descripcion': 'Crear un reporte de falla'},
        {'id': 6, 'titulo': '¿Como creamos una orden de servicio?', 'categoria': 'Softv', 'imagenes': ['softv/softv23.png', 'softv/softv24.png', 'softv/softv26.png', 'softv/softv27.png', 'softv/softv28.png'], 'descripcion': 'Crear una orden de servicio'},
        {'id': 7, 'titulo': '¿Como borramos un reporte de falla en caso necesario?', 'categoria': 'Softv', 'imagenes': ['softv/softv29.png', 'softv/softv29.png', 'softv/softv29.png'], 'descripcion': 'Como eliminar un reporte de falla'},
        {'id': 8, 'titulo': '¿Como ingresamos un nuevo cliente?', 'categoria': 'Softv', 'imagenes': ['softv/softv30.png', 'softv/softv31.png', 'softv/softv32.png', 'softv/softv33.png', 'softv/softv32.png'], 'descripcion': 'Crear un nuevo cliente'},
        {'id': 9, 'titulo': '¿Como buscar un usuario?', 'categoria': 'Vortex', 'imagenes': ['vortex/vortex1.png', 'vortex/vortex2.png', 'vortex/vortex3.png'], 'descripcion': 'Buscar a un usuario'},
        {'id': 10, 'titulo': '¿Como validar puertos en uso y la MAC del equipo?', 'categoria': 'Vortex', 'imagenes': ['vortex/vortex4.png', 'vortex/vortex5.png'], 'descripcion': 'Como validar si el usuario esta haciendo uso de los puertos o el dispositivo no da MAC'},
        {'id': 11, 'titulo': '¿Como validar si el usuario esta teniendo consumo del servicio?', 'categoria': 'Vortex', 'imagenes': ['vortex/vortex7.png'], 'descripcion': 'Como validar el consumo del usuario'},
        {'id': 12, 'titulo': '¿Como cambiar la VLAN?', 'categoria': 'Vortex', 'imagenes': ['vortex/vortex8.png', 'vortex/vortex9.png'], 'descripcion': 'Como cambiar la VLAN acorde a la zona'},
        {'id': 13, 'titulo': '¿Como realizar un resync config?', 'categoria': 'Vortex', 'imagenes': ['vortex/vortex10.png', 'vortex/vortex11.png'], 'descripcion': 'Como realizar un resync config'},
        {'id': 14, 'titulo': '¿Como realizar un reboot?', 'categoria': 'Vortex', 'imagenes': ['vortex/vortex12.png', 'vortex/vortex13.png'], 'descripcion': 'Como realizar un reebot'},
        {'id': 15, 'titulo': '¿Como identificar si el servicio de internet y TV estan activados?', 'categoria': 'Vortex', 'imagenes': ['vortex/vortex14.png'], 'descripcion': 'Validar si el servicio esta activo'},
    ]
    return render_template('soluciones_visuales.html', soluciones=soluciones)

@app.route('/atencion_telefonica')
@login_required
def atencion_telefonica():
    return render_template('atencion_telefonica.html')

@app.route('/informacion-general')
@login_required
def informacion_general():
    informacion = {
        'planes': {
            'titulo': '📡 Planes de Servicio', 'icono': 'fa-tv',
            'contenido': [
                {'subtitulo': 'Planes Básicos', 'contenido_items': ['💯 *PLANES DE TV E INTERNET* 💯', '400 megas + TV: $85.000', '500 megas + TV: $95.000', '600 megas + TV: $105.000', '', '💯 *PLANES SOLO TV* 💯', '10Mb + TV: $50.000', '', '🌐 *PLANES SOLO INTERNET* 🌐', '400 megas: $75.000', '500 megas: $85.000', '600 megas: $95.000']},
                {'subtitulo': 'Planes Corporativos', 'contenido_items': ['💯 *PLANES CORPORATIVOS* 💯', '1Mb: $12.000', '30Mb (mínimo): $360.000 + 19% IVA = $428.400', '*Planes hogar:* se agrega 19% IVA', '*Equipo:* robusto para configuraciones especiales']},
                {'subtitulo': 'Planes Guamal y Sanmartin', 'contenido_items': ['🎯 PLANES DE TV + INTERNET 🎯', 'TV + 200MB: $65.000', 'TV + 300MB: $75.000', 'TV + 400MB: $85.000', '', '📺 PLAN SOLO TV 📺', 'Solo TV: $50.000']},
                {'subtitulo': 'Planes Acacías', 'contenido_items': ['💯 *PLANES DE TV E INTERNET* 💯', 'TV + Internet 200MB: $85.000', 'TV + Internet 300MB: $95.000', 'TV + Internet 400MB: $105.000', '', '💯 *PLANES SOLO TV* 💯', 'Solo TV: $50.000', '', '🌐 *PLANES SOLO INTERNET* 🌐', '200MB: $75.000', '300MB: $85.000', '400MB: $95.000']},
            ]
        },
        'afiliaciones': {
            'titulo': '👥 Afiliaciones', 'icono': 'fa-user-plus',
            'contenido': [
                {'subtitulo': 'Información General para Afiliar', 'contenido_items': ['*La afiliación no tiene costo*', '*Instalación sin costo* en zona urbana (rural: $150.000)', '', '*Requisitos:*', '• 1 Fotocopia de la cédula', '• 1 Fotocopia del recibo de agua o luz', '• Pago del primer mes por anticipado', '• Servicio de TV para 2 televisores', '', '*Puntos adicionales de TV:*', '• Cada punto: $20.000 (solo instalación)', '• Mensualidad no cambia', '• Solo para el mismo predio', '', '*Señal Digital:*', '• Decodificador: $58.000 (único pago)', '• Para TVs clásicos con señal analógica', '', '*Tiempo de instalación:* 2-4 días hábiles']},
                {'subtitulo': 'Afiliación San Joaquín', 'contenido_items': ['*Costo de instalación:* $60.000', '*Fibra incluida:* primeros 70 metros', '*Costo metro adicional:* $1.700', '', '*Servicio de TV:* 1 televisor', '*Puntos adicionales:* $35.000 c/u', '*Requisitos y tiempos iguales*  a afiliación general']},
                {'subtitulo': 'Información Adicional', 'contenido_items': ['*Para asesores solicitar:*', '• Barrio', '• Dirección exacta', '• Nombre del titular', '• 2 números de teléfono', '', '*Sin cláusula de permanencia*', '*Pago por adelantado* después de firmar contrato', '*Contrato*  se envía y recibe por el mismo medio']},
            ]
        },
        'win_sports': {
            'titulo': '⚽ Win Sports +', 'icono': 'fa-futbol',
            'contenido': [{'subtitulo': '¡Llegó Win Sports + a M@STV Producciones!', 'contenido_items': ['*Precio:* $35.000 adicionales al mes', '*Incluye:*', '• Acceso a Win Sports +', '• 14 canales premium', '• Y mucho más contenido deportivo', '', '*TV Box:* $100.000 (costo único)', '*No necesario* si TV es Android (con Google Play Store)', '*Cláusula:* 6 meses', '*Requisito:* Tener plan de internet con nosotros']}]
        },
        'oficinas': {
            'titulo': '🏢 Oficinas y Horarios', 'icono': 'fa-building',
            'contenido': [
                {'subtitulo': 'Horarios de Atención', 'contenido_items': ['*Lunes a Viernes:* 8:00 AM - 5:00 PM', '*Sábados:* 8:00 AM - 12:00 PM']},
                {'subtitulo': 'Direcciones de Oficinas', 'contenido_items': ['*Facatativá:* Cl 11 #7A-04, Diurba', '*Bojacá:* Cr 6 #5-146, Barrio Centro', '*Zipacón:* Crr 4 #5-57, Frente al parque', '*Rosal:* Cr 8 #8-08, Local 3 Centro', '*El Triunfo:* Crr 3 #2-40, Frente al coliseo', '*Viotá:* Cl 20 #11-10, Frente a estación de policía', '*Girardot:* Crr 10 #18-44, Barrio Centro / Frente a Bancamía', '*Cachipay:* Crr 3 #3-36, Barrio Centro', '*Sasaima:* Crr 2 #3-30, Barrio 3 Esquinas', '*La Mesa:* Cl 8 #16-59, Barrio Santa Bárbara', '*Anolaima:* Crr 7 #02-57, Barrio Centro', '*Mesitas del Colegio:* Cl 10 #6-37, Barrio Centro', '*Anapoima:* Cr 2 #7-32, Local 2 Centro', '*Albán:* Cl 4 #2-04, Punto de Servientrega', '*Madrid:* Cl 12 #3-64, Barrio Arrayane', '*Guayabal de Síquima:* Cl 3 #5-28', '*Tocaima:* Cl 4 #9-75', '*San Joaquín:* Cr 4 N 4-55, Al lado del árbol de los aburridos', '*Apulo:* Cl 14 #6-23, Local 102', '*Villeta:* Cr 5 #3-43, Local 6 Torre 4 Conjunto Santa Cruz', '*Acacías:* Cl 15 #22-40, Local 12, Edificio Dark Gym', '*San Martín:* Cl 7 #5-34, Barrio Fundadores', '*Guamal:* Cl 10 #4A-04, Barrio Las Villas', '*Quipile:* Crr 2 #6-07']},
                {'subtitulo': 'Puntos Autorizados Facatativá', 'contenido_items': ['*Bolos el Tunjo:* Cr 2 #6-105', '*CLT Comunicaciones:* Cl 19 #1A-28 Sur, Prado de Cartagenita', '*Portal de María:* Transversal 11 #5-04, Manzana 5 Casa 30 S.M.A.', '*Papelería Expresate:* Cl 8 #10-05, Zambrano', '*One Books:* Diagonal 5 Este #9E-02, Juan Pablo II', '*Papelería Chico 1:* Cr 3 #5B-08 Este, Chico 1']},
            ]
        },
        'procesos': {
            'titulo': '📋 Procesos y Trámites', 'icono': 'fa-clipboard-list',
            'contenido': [
                {'subtitulo': 'Cancelación de Servicio', 'contenido_items': ['*Requisitos:*', '• Acercarse a la oficina', '• Carta indicando razón de cancelación', '• Paz y salvo', '• Equipos instalados (equipos y cargadores)']},
                {'subtitulo': 'Cambio de Titular', 'contenido_items': ['*Requisitos:*', '• Carta solicitando cambio, firmada por antiguo y nuevo titular', '• Copia de cédula del nuevo titular', '• Estar al día en los pagos']},
                {'subtitulo': 'Cambio de Plan', 'contenido_items': ['*Procedimiento:*', '• Acercarse a la oficina', '• Carta solicitando cambio de plan', '• Estar al día en pagos', '• Cancelar por adelantado valor del nuevo plan', '• Ideal realizarlo a finales de mes']},
                {'subtitulo': 'Traslado de Domicilio', 'contenido_items': ['*Costo:* $20.000', '*Puntos adicionales:* $10.000 c/u (movimiento)', '*Tiempo:* 2-3 días hábiles', '*Requisito:* Llevar equipos a la nueva residencia']},
                {'subtitulo': 'Solicitud de Facturas', 'contenido_items': ['*Datos requeridos:*', '• Contrato', '• Nombre completo', '• Cédula', '• Correo electrónico', '• Teléfono', '• Dirección completa', '• Municipio y barrio', '• Plan de internet', '• Valor del plan', '• Estrato', '*Empresas:* enviar foto del RUT']},
            ]
        },
        'contacto': {
            'titulo': '📞 Contacto y Soporte', 'icono': 'fa-headset',
            'contenido': [{'subtitulo': 'Información de Contacto', 'contenido_items': ['*Email PQR:* pqr@mastvproducciones.net.co', '*Email CARTERA:* auxiliaradministrativo@mastvproducciones.net.co', '*Email INGENIERIA:* ingenieria@mastvproducciones.net.co', '*Email RECURSOS HUMANOS:* rh@mastvproducciones.net.co', '*Chat de Soporte:* Solo mensajes escritos 3187777771', '*No se reciben:* audios ni llamadas por WhatsApp']}]
        }
    }
    return render_template('informacion_general.html', informacion=informacion)

# ===== RUTAS SST =====
@app.route('/sst')
@login_required
def sst_dashboard():
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    return render_template('sst/dashboard.html')

@app.route('/sst/contenido')
@login_required
def sst_contenido():
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    contenido = []
    categorias = []
    try:
        categorias_data = obtener_categorias_sst()
        for cat in categorias_data:
            categorias.append({'id': cat[0], 'nombre': cat[1], 'color': cat[2]})
        filtros = {'query': request.args.get('q', ''), 'categoria': request.args.get('categoria', ''), 'tipo': request.args.get('tipo', '')}
        contenido_data = obtener_contenido_sst(filtros)
        for item in contenido_data:
            tags_value = item[12]
            tags_str = '' if tags_value is None else str(tags_value)
            contenido_dict = {
                'id': item[0], 'titulo': item[1], 'descripcion': item[2], 'tipo': item[3],
                'archivo_url': item[4], 'tiene_archivo': item[5] is not None,
                'archivo_nombre': item[6], 'archivo_tipo': item[7], 'archivo_tamano': item[8],
                'video_url': item[9], 'categoria_id': item[10], 'es_obligatorio': item[11],
                'tags': tags_str, 'fecha_publicacion': item[13], 'usuario_creador': item[14],
                'categoria_nombre': item[15], 'categoria_color': item[16], 'creador_nombre': item[17]
            }
            contenido.append(contenido_dict)
    except Exception as e:
        flash('Error al cargar el contenido SST', 'error')
        logger.error(f"❌ Error en sst_contenido: {e}")
    return render_template('sst/contenido.html', contenido=contenido, categorias=categorias)

@app.route('/sst/agregar', methods=['GET', 'POST'])
@login_required
@retry_on_ssl_error(max_retries=2, delay=3)
def sst_agregar_contenido():
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    if current_user.rol not in ['admin', 'sst']:
        flash('No tienes permisos para agregar contenido SST', 'error')
        return redirect(url_for('sst_dashboard'))
    categorias = []
    try:
        categorias_data = obtener_categorias_sst()
        for cat in categorias_data:
            categorias.append({'id': cat[0], 'nombre': cat[1], 'color': cat[2]})
        if request.method == 'POST':
            titulo = request.form.get('titulo', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            tipo = request.form.get('tipo', '').strip()
            categoria_id = request.form.get('categoria_id', '').strip()
            es_obligatorio = 'es_obligatorio' in request.form
            tags = request.form.get('tags', '').strip()
            video_url = request.form.get('video_url', '').strip()
            archivo_url = request.form.get('archivo_url', '').strip()
            if not titulo or not tipo or not categoria_id:
                flash('❌ Todos los campos obligatorios deben ser completados', 'error')
                return render_template('sst/agregar_contenido.html', categorias=categorias)
            try:
                categoria_id_int = int(categoria_id)
            except (ValueError, TypeError):
                flash('❌ Categoría inválida', 'error')
                return render_template('sst/agregar_contenido.html', categorias=categorias)
            archivo_data = None
            file = request.files.get('archivo_local')
            if file and file.filename != '':
                if allowed_file(file.filename):
                    file.seek(0, 2)
                    file_size = file.tell()
                    file.seek(0)
                    logger.info(f"📦 Procesando archivo: {file.filename} ({file_size} bytes)")
                    if file_size > 5 * 1024 * 1024:
                        chunks = []
                        while True:
                            chunk = file.read(8192)
                            if not chunk:
                                break
                            chunks.append(chunk)
                        file_data = b''.join(chunks)
                        file_name = generar_nombre_seguro(file.filename)
                        file_type = mimetypes.guess_type(file_name)[0] or 'application/octet-stream'
                        archivo_data = {'data': file_data, 'nombre': file_name, 'tipo': file_type, 'tamano': len(file_data)}
                    else:
                        archivo_data = guardar_archivo_en_bd(file)
                    if not archivo_data:
                        flash('❌ Error al procesar el archivo', 'error')
                        return render_template('sst/agregar_contenido.html', categorias=categorias)
                    video_url = None
                    archivo_url = None
                else:
                    extensiones_permitidas = ', '.join(app.config['ALLOWED_EXTENSIONS'])
                    flash(f'❌ Tipo de archivo no permitido. Extensiones válidas: {extensiones_permitidas}', 'error')
                    return render_template('sst/agregar_contenido.html', categorias=categorias)
            validation_error = None
            if tipo == 'video':
                if not video_url and not archivo_data:
                    validation_error = 'Para video debe proporcionar una URL de video o subir un archivo'
            elif tipo in ['documento', 'imagen']:
                if not archivo_url and not archivo_data:
                    validation_error = 'Debe proporcionar una URL o subir un archivo'
            elif tipo == 'enlace':
                if not archivo_url:
                    validation_error = 'Debe proporcionar una URL para enlaces'
                archivo_data = None
                video_url = None
            if validation_error:
                flash(f'❌ {validation_error}', 'error')
                return render_template('sst/agregar_contenido.html', categorias=categorias)
            video_url = video_url if video_url else None
            archivo_url = archivo_url if archivo_url else None
            descripcion = descripcion if descripcion else None
            tags = tags if tags else None
            try:
                success = insertar_contenido_con_archivo(
                    titulo=titulo, descripcion=descripcion, tipo=tipo,
                    categoria_id=categoria_id_int, es_obligatorio=es_obligatorio,
                    tags=tags, usuario_creador=current_user.id, archivo_data=archivo_data,
                    video_url=video_url, archivo_url=archivo_url)
                if success:
                    flash('✅ Contenido SST agregado correctamente', 'success')
                    return redirect(url_for('sst_contenido'))
                else:
                    flash('❌ Error al guardar en la base de datos', 'error')
            except Exception as db_error:
                flash(f'❌ Error de base de datos: {str(db_error)}', 'error')
                return render_template('sst/agregar_contenido.html', categorias=categorias)
    except Exception as e:
        flash(f'❌ Error al agregar contenido SST: {str(e)}', 'error')
        logger.error(f"❌ ERROR GENERAL EN SST_AGREGAR_CONTENIDO: {e}")
    return render_template('sst/agregar_contenido.html', categorias=categorias)

@app.route('/sst/archivo/<int:id>')
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_descargar_archivo(id):
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    try:
        archivo = obtener_archivo_desde_bd(id)
        if not archivo:
            flash('Archivo no encontrado', 'error')
            return redirect(url_for('sst_contenido'))
        if not archivo.get('data'):
            flash('El archivo está vacío', 'error')
            return redirect(url_for('sst_contenido'))
        file_data = BytesIO(archivo['data'])
        return send_file(file_data, mimetype=archivo['tipo'], as_attachment=False, download_name=archivo['nombre'])
    except Exception as e:
        flash(f'Error al descargar el archivo: {str(e)}', 'error')
        logger.error(f"❌ Error en sst_descargar_archivo: {e}")
        return redirect(url_for('sst_contenido'))

@app.route('/sst/archivo/descargar/<int:id>')
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_descargar_archivo_forzado(id):
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    try:
        archivo = obtener_archivo_desde_bd(id)
        if not archivo or not archivo.get('data'):
            flash('Archivo no encontrado', 'error')
            return redirect(url_for('sst_contenido'))
        file_data = BytesIO(archivo['data'])
        return send_file(file_data, mimetype=archivo['tipo'], as_attachment=True, download_name=archivo['nombre'])
    except Exception as e:
        flash(f'Error al descargar el archivo: {str(e)}', 'error')
        logger.error(f"❌ Error en sst_descargar_archivo_forzado: {e}")
        return redirect(url_for('sst_contenido'))

@app.route('/sst/contenido/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_editar_contenido(id):
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    if current_user.rol not in ['admin', 'sst']:
        flash('No tienes permisos para editar contenido SST', 'error')
        return redirect(url_for('sst_dashboard'))
    contenido = None
    categorias = []
    try:
        categorias_data = obtener_categorias_sst()
        for cat in categorias_data:
            categorias.append({'id': cat[0], 'nombre': cat[1], 'color': cat[2]})
        if request.method == 'POST':
            titulo = request.form.get('titulo', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            tipo = request.form.get('tipo', '').strip()
            categoria_id = request.form.get('categoria_id', '').strip()
            es_obligatorio = 'es_obligatorio' in request.form
            tags = request.form.get('tags', '').strip()
            video_url = request.form.get('video_url', '').strip() or None
            archivo_url = request.form.get('archivo_url', '').strip() or None
            if not titulo or not tipo or not categoria_id:
                flash('Todos los campos obligatorios deben ser completados', 'error')
                return render_template('sst/editar_contenido.html', contenido=contenido, categorias=categorias)
            archivo_data = None
            file = request.files.get('archivo_local')
            if file and file.filename != '':
                if allowed_file(file.filename):
                    archivo_data = guardar_archivo_en_bd(file)
                    if not archivo_data:
                        flash('Error al procesar el archivo', 'error')
                        return render_template('sst/editar_contenido.html', contenido=contenido, categorias=categorias)
                    video_url = None
                    archivo_url = None
                else:
                    flash('Tipo de archivo no permitido', 'error')
                    return render_template('sst/editar_contenido.html', contenido=contenido, categorias=categorias)
            if archivo_data:
                ejecutar_consulta("""
                    UPDATE sst_contenido
                    SET titulo=%s, descripcion=%s, tipo=%s, archivo_url=%s,
                        archivo_data=%s, archivo_nombre=%s, archivo_tipo=%s, archivo_tamano=%s,
                        video_url=%s, categoria_id=%s, es_obligatorio=%s,
                        tags=%s, fecha_actualizacion=CURRENT_TIMESTAMP
                    WHERE id=%s
                """, (titulo, descripcion, tipo, archivo_url,
                      psycopg2.Binary(archivo_data['data']), archivo_data['nombre'],
                      archivo_data['tipo'], archivo_data['tamano'],
                      video_url, categoria_id, es_obligatorio, tags, id), commit=True)
            else:
                ejecutar_consulta("""
                    UPDATE sst_contenido
                    SET titulo=%s, descripcion=%s, tipo=%s, archivo_url=%s,
                        video_url=%s, categoria_id=%s, es_obligatorio=%s,
                        tags=%s, fecha_actualizacion=CURRENT_TIMESTAMP
                    WHERE id=%s
                """, (titulo, descripcion, tipo, archivo_url, video_url,
                      categoria_id, es_obligatorio, tags, id), commit=True)
            flash('✅ Contenido actualizado correctamente', 'success')
            return redirect(url_for('sst_contenido'))
        resultado = ejecutar_consulta("""
            SELECT sc.*, cat.nombre as categoria_nombre, cat.color as categoria_color,
                   u.usuario as creador_nombre
            FROM sst_contenido sc
            LEFT JOIN sst_categorias cat ON sc.categoria_id = cat.id
            LEFT JOIN usuarios u ON sc.usuario_creador = u.id
            WHERE sc.id = %s
        """, (id,), fetch=True)
        if resultado and resultado[0]:
            contenido_data = resultado[0]
            contenido = {
                'id': contenido_data[0], 'titulo': contenido_data[1], 'descripcion': contenido_data[2],
                'tipo': contenido_data[3], 'archivo_url': contenido_data[4],
                'tiene_archivo': contenido_data[5] is not None, 'archivo_nombre': contenido_data[6],
                'archivo_tipo': contenido_data[7], 'archivo_tamano': contenido_data[8],
                'video_url': contenido_data[9], 'categoria_id': contenido_data[10],
                'es_obligatorio': contenido_data[11],
                'tags': str(contenido_data[12]) if contenido_data[12] is not None else '',
                'fecha_publicacion': contenido_data[13], 'usuario_creador': contenido_data[14],
                'categoria_nombre': contenido_data[15], 'categoria_color': contenido_data[16],
                'creador_nombre': contenido_data[17]
            }
    except Exception as e:
        flash(f'Error al editar contenido SST: {str(e)}', 'error')
        logger.error(f"❌ Error en sst_editar_contenido: {e}")
    if not contenido:
        flash('Contenido no encontrado', 'error')
        return redirect(url_for('sst_contenido'))
    return render_template('sst/editar_contenido.html', contenido=contenido, categorias=categorias)

@app.route('/sst/contenido/<int:id>/eliminar', methods=['POST'])
@login_required
def sst_eliminar_contenido(id):
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    if current_user.rol not in ['admin', 'sst']:
        flash('No tienes permisos para eliminar contenido SST', 'error')
        return redirect(url_for('sst_contenido'))
    try:
        ejecutar_consulta("DELETE FROM sst_contenido WHERE id = %s", (id,), commit=True)
        flash('✅ Contenido eliminado correctamente', 'success')
    except Exception as e:
        flash(f'Error al eliminar contenido SST: {str(e)}', 'error')
        logger.error(f"❌ Error en sst_eliminar_contenido: {e}")
    return redirect(url_for('sst_contenido'))

@app.route('/sst/video/<int:id>')
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_ver_video(id):
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    video = None
    try:
        resultado = ejecutar_consulta("""
            SELECT sc.*, cat.nombre as categoria_nombre, cat.color as categoria_color
            FROM sst_contenido sc
            LEFT JOIN sst_categorias cat ON sc.categoria_id = cat.id
            WHERE sc.id = %s
        """, (id,), fetch=True)
        if resultado and resultado[0]:
            video_data = resultado[0]
            video = {
                'id': video_data[0], 'titulo': video_data[1], 'descripcion': video_data[2],
                'tipo': video_data[3], 'archivo_nombre': video_data[6], 'archivo_tipo': video_data[7],
                'archivo_tamano': video_data[8], 'video_url': video_data[9],
                'tiene_archivo': video_data[5] is not None,
                'categoria_nombre': video_data[15] if len(video_data) > 15 else '',
                'categoria_color': video_data[16] if len(video_data) > 16 else '#007bff',
                'fecha_publicacion': video_data[13], 'es_obligatorio': video_data[11]
            }
        else:
            flash('Video no encontrado', 'error')
            return redirect(url_for('sst_contenido'))
    except Exception as e:
        flash(f'Error al cargar el video: {str(e)}', 'error')
        logger.error(f"❌ Error en sst_ver_video: {e}")
        return redirect(url_for('sst_contenido'))
    if not video:
        flash('Video no encontrado', 'error')
        return redirect(url_for('sst_contenido'))
    return render_template('sst/ver_video.html', video=video)

@app.route('/sst/video/stream/<int:id>')
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_stream_video(id):
    if not current_user.puede('acceder_sst'):
        return Response('No autorizado', status=403)
    try:
        archivo = obtener_archivo_desde_bd(id)
        if not archivo or not archivo.get('data'):
            return Response('Video no encontrado', status=404)
        if not archivo['tipo'].startswith('video/'):
            return Response('El archivo no es un video', status=400)
        file_data = BytesIO(archivo['data'])
        return send_file(file_data, mimetype=archivo['tipo'], as_attachment=False)
    except Exception as e:
        logger.error(f"❌ Error en sst_stream_video: {e}")
        return Response('Error interno del servidor', status=500)

# ===== RUTAS PARA SERVIR ARCHIVOS ESTÁTICOS =====
@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory('static', filename)

# ===== API PARA PROBLEMAS =====
@app.route('/api/problemas/<categoria>')
@login_required
def obtener_problemas(categoria):
    if not current_user.puede('acceder_soporte'):
        return jsonify([])
    problemas_por_categoria = {
        'TV': ['No hay señal en el televisor', 'Imagen pixelada o con interferencias', 'Sin sonido en algunos canales', 'Problemas con la guía de programación', 'Otro problema con TV'],
        'Internet': ['Internet lento o intermitente', 'Sin conexión a internet', 'Problemas con WiFi', 'No puedo conectarme a sitios específicos', 'Velocidad inferior a la contratada', 'Problemas con el módem/router', 'Otro problema con Internet'],
        'Equipo': ['Equipo no enciende', 'Problemas con puertos HDMI/USB', 'Dispositivo no da MAC', 'Problemas niveles opticos', 'Otro problema con Equipo']
    }
    return jsonify(problemas_por_categoria.get(categoria, []))


# ===== RUTAS PARA GESTIÓN DEL PLAN ANUAL DE TRABAJO PESV =====
@app.route('/sst/plan-anual')
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_plan_anual():
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COUNT(*) as total,
                SUM(CASE WHEN estado = 'completado' THEN 1 ELSE 0 END) as completadas,
                SUM(CASE WHEN estado = 'en_proceso' THEN 1 ELSE 0 END) as en_proceso,
                SUM(CASE WHEN estado = 'pendiente' THEN 1 ELSE 0 END) as pendientes,
                ROUND(AVG(porcentaje_avance), 2) as promedio_avance
            FROM plan_anual_trabajo
        """)
        stats = cursor.fetchone()
        cursor.execute("""
            SELECT ciclo_phva, COUNT(*) as total,
                SUM(CASE WHEN estado = 'completado' THEN 1 ELSE 0 END) as completadas,
                ROUND(AVG(porcentaje_avance), 2) as promedio
            FROM plan_anual_trabajo WHERE ciclo_phva IS NOT NULL
            GROUP BY ciclo_phva
            ORDER BY CASE ciclo_phva WHEN 'Planear' THEN 1 WHEN 'Hacer' THEN 2 WHEN 'Verificar' THEN 3 WHEN 'Actuar' THEN 4 ELSE 5 END
        """)
        stats_phva = cursor.fetchall()
        cursor.execute("""
            SELECT id, actividad, ciclo_phva, responsables, estado, porcentaje_avance, fecha_actualizacion
            FROM plan_anual_trabajo ORDER BY fecha_actualizacion DESC LIMIT 10
        """)
        actividades_recientes = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('sst/plan_anual_dashboard.html', stats=stats, stats_phva=stats_phva, actividades_recientes=actividades_recientes)
    except Exception as e:
        flash(f'Error al cargar el plan anual: {str(e)}', 'error')
        logger.error(f"Error en sst_plan_anual: {e}")
        return redirect(url_for('sst_dashboard'))

@app.route('/sst/plan-anual/actividades')
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_plan_anual_actividades():
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    ciclo = request.args.get('ciclo', '')
    estado = request.args.get('estado', '')
    responsable = request.args.get('responsable', '')
    mes = request.args.get('mes', '')
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        query = "SELECT id, actividad, evidencia, ciclo_phva, responsables, estado, porcentaje_avance, nivel_pesv FROM plan_anual_trabajo WHERE 1=1"
        params = []
        if ciclo:
            query += " AND ciclo_phva = %s"
            params.append(ciclo)
        if estado:
            query += " AND estado = %s"
            params.append(estado)
        if responsable:
            query += " AND responsables ILIKE %s"
            params.append(f'%{responsable}%')
        if mes:
            meses = {'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8, 'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12}
            mes_lower = mes.lower()
            if mes_lower in meses:
                query += f" AND ({mes_lower}_semana1_p = TRUE OR {mes_lower}_semana2_p = TRUE OR {mes_lower}_semana3_p = TRUE OR {mes_lower}_semana4_p = TRUE)"
        query += " ORDER BY ciclo_phva, actividad"
        cursor.execute(query, params)
        actividades = cursor.fetchall()
        cursor.execute("SELECT DISTINCT ciclo_phva FROM plan_anual_trabajo WHERE ciclo_phva IS NOT NULL ORDER BY ciclo_phva")
        ciclos_disponibles = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT responsables FROM plan_anual_trabajo WHERE responsables IS NOT NULL")
        responsables_disponibles = set()
        for row in cursor.fetchall():
            if row[0]:
                for r in row[0].split('-'):
                    responsables_disponibles.add(r.strip())
        responsables_disponibles = sorted(list(responsables_disponibles))
        cursor.close()
        conn.close()
        return render_template('sst/plan_anual_actividades.html', actividades=actividades,
                             ciclos=ciclos_disponibles, responsables_list=responsables_disponibles,
                             filtro_ciclo=ciclo, filtro_estado=estado, filtro_responsable=responsable, filtro_mes=mes)
    except Exception as e:
        flash(f'Error al cargar actividades: {str(e)}', 'error')
        logger.error(f"Error en sst_plan_anual_actividades: {e}")
        return redirect(url_for('sst_plan_anual'))

@app.route('/sst/plan-anual/actividad/<int:id>')
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_plan_anual_actividad_detalle(id):
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para ver esta actividad', 'error')
        return redirect(url_for('sst_plan_anual'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, actividad, evidencia, ciclo_phva, articulos_decreto,
                nivel_pesv, responsables, recursos,
                enero_semana1_p, enero_semana1_e, enero_semana2_p, enero_semana2_e,
                enero_semana3_p, enero_semana3_e, enero_semana4_p, enero_semana4_e,
                febrero_semana1_p, febrero_semana1_e, febrero_semana2_p, febrero_semana2_e,
                febrero_semana3_p, febrero_semana3_e, febrero_semana4_p, febrero_semana4_e,
                marzo_semana1_p, marzo_semana1_e, marzo_semana2_p, marzo_semana2_e,
                marzo_semana3_p, marzo_semana3_e, marzo_semana4_p, marzo_semana4_e,
                abril_semana1_p, abril_semana1_e, abril_semana2_p, abril_semana2_e,
                abril_semana3_p, abril_semana3_e, abril_semana4_p, abril_semana4_e,
                mayo_semana1_p, mayo_semana1_e, mayo_semana2_p, mayo_semana2_e,
                mayo_semana3_p, mayo_semana3_e, mayo_semana4_p, mayo_semana4_e,
                junio_semana1_p, junio_semana1_e, junio_semana2_p, junio_semana2_e,
                junio_semana3_p, junio_semana3_e, junio_semana4_p, junio_semana4_e,
                julio_semana1_p, julio_semana1_e, julio_semana2_p, julio_semana2_e,
                julio_semana3_p, julio_semana3_e, julio_semana4_p, julio_semana4_e,
                agosto_semana1_p, agosto_semana1_e, agosto_semana2_p, agosto_semana2_e,
                agosto_semana3_p, agosto_semana3_e, agosto_semana4_p, agosto_semana4_e,
                septiembre_semana1_p, septiembre_semana1_e, septiembre_semana2_p, septiembre_semana2_e,
                septiembre_semana3_p, septiembre_semana3_e, septiembre_semana4_p, septiembre_semana4_e,
                octubre_semana1_p, octubre_semana1_e, octubre_semana2_p, octubre_semana2_e,
                octubre_semana3_p, octubre_semana3_e, octubre_semana4_p, octubre_semana4_e,
                noviembre_semana1_p, noviembre_semana1_e, noviembre_semana2_p, noviembre_semana2_e,
                noviembre_semana3_p, noviembre_semana3_e, noviembre_semana4_p, noviembre_semana4_e,
                diciembre_semana1_p, diciembre_semana1_e, diciembre_semana2_p, diciembre_semana2_e,
                diciembre_semana3_p, diciembre_semana3_e, diciembre_semana4_p, diciembre_semana4_e,
                observaciones, estado, porcentaje_avance, fecha_creacion, fecha_actualizacion, usuario_actualizacion
            FROM plan_anual_trabajo WHERE id = %s
        """, (id,))
        actividad_raw = cursor.fetchone()
        if not actividad_raw:
            flash('❌ Actividad no encontrada', 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('sst_plan_anual_actividades'))
        actividad = {
            'id': actividad_raw[0], 'actividad': actividad_raw[1], 'evidencia': actividad_raw[2],
            'ciclo_phva': actividad_raw[3], 'articulos_decreto': actividad_raw[4],
            'nivel_pesv': actividad_raw[5], 'responsables': actividad_raw[6], 'recursos': actividad_raw[7],
        }
        meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
        programacion = {}
        semanas_planificadas = 0
        semanas_ejecutadas = 0
        col_idx = 8
        for mes in meses:
            programacion[mes] = []
            for semana in range(1, 5):
                planificado = actividad_raw[col_idx] if actividad_raw[col_idx] else False
                ejecutado = actividad_raw[col_idx + 1] if actividad_raw[col_idx + 1] else False
                if planificado:
                    semanas_planificadas += 1
                if ejecutado:
                    semanas_ejecutadas += 1
                programacion[mes].append({'semana': semana, 'planificado': planificado, 'ejecutado': ejecutado})
                col_idx += 2
        actividad['observaciones'] = actividad_raw[104] if len(actividad_raw) > 104 else ''
        actividad['estado'] = actividad_raw[105] if len(actividad_raw) > 105 else 'pendiente'
        try:
            porcentaje_raw = actividad_raw[106] if len(actividad_raw) > 106 else 0
            actividad['porcentaje_avance'] = 0.0 if (porcentaje_raw is None or porcentaje_raw == '') else float(porcentaje_raw)
        except (ValueError, TypeError):
            actividad['porcentaje_avance'] = 0.0
        actividad['fecha_creacion'] = actividad_raw[107] if len(actividad_raw) > 107 else None
        actividad['fecha_actualizacion'] = actividad_raw[108] if len(actividad_raw) > 108 else None
        actividad['usuario_actualizacion'] = actividad_raw[109] if len(actividad_raw) > 109 else None
        actividad['programacion'] = programacion
        actividad['semanas_planificadas'] = semanas_planificadas
        actividad['semanas_ejecutadas'] = semanas_ejecutadas
        try:
            cursor.execute("SELECT id, titulo, descripcion, archivo_nombre, fecha_carga FROM plan_evidencias WHERE plan_id = %s ORDER BY fecha_carga DESC", (id,))
            evidencias = cursor.fetchall()
        except Exception as e:
            logger.error(f"❌ Error cargando evidencias: {e}")
            evidencias = []
        try:
            cursor.execute("""
                SELECT s.id, s.comentario, s.tipo, s.fecha, u.usuario
                FROM plan_seguimiento s LEFT JOIN usuarios u ON s.usuario_id = u.id
                WHERE s.actividad_id = %s ORDER BY s.fecha DESC
            """, (id,))
            seguimientos = cursor.fetchall()
        except Exception as e:
            logger.warning(f"No se pudieron cargar seguimientos: {e}")
            seguimientos = []
        cursor.close()
        conn.close()
        return render_template('sst/plan_anual_detalle.html', actividad=actividad, evidencias=evidencias, seguimientos=seguimientos)
    except Exception as e:
        flash(f'❌ Error al cargar detalle: {str(e)}', 'error')
        logger.error(f"Error en sst_plan_anual_actividad_detalle: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return redirect(url_for('sst_plan_anual_actividades'))

@app.route('/sst/plan-anual/actividad/<int:id>/actualizar', methods=['POST'])
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_plan_anual_actualizar_actividad(id):
    if not current_user.puede('gestionar_plan_anual'):
        flash('No tienes permisos para modificar el plan anual', 'error')
        return redirect(url_for('sst_plan_anual_actividad_detalle', id=id))
    try:
        mes = request.form.get('mes')
        semana = request.form.get('semana')
        ejecutado = request.form.get('ejecutado') == 'true'
        if not mes or not semana:
            flash('Datos incompletos', 'error')
            return redirect(url_for('sst_plan_anual_actividad_detalle', id=id))
        conn = crear_conexion()
        cursor = conn.cursor()
        columna = f"{mes}_semana{semana}_e"
        cursor.execute(f"UPDATE plan_anual_trabajo SET {columna} = %s, fecha_actualizacion = CURRENT_TIMESTAMP, usuario_actualizacion = %s WHERE id = %s",
                      (ejecutado, current_user.id, id))
        nuevo_estado = 'en_proceso' if ejecutado else 'pendiente'
        cursor.execute("UPDATE plan_anual_trabajo SET estado = %s WHERE id = %s", (nuevo_estado, id))
        conn.commit()
        cursor.close()
        conn.close()
        flash('Actividad actualizada exitosamente', 'success')
    except Exception as e:
        flash(f'Error al actualizar: {str(e)}', 'error')
        logger.error(f"Error en sst_plan_anual_actualizar_actividad: {e}")
    return redirect(url_for('sst_plan_anual_actividad_detalle', id=id))

@app.route('/sst/plan-anual/actividad/<int:id>/evidencia', methods=['POST'])
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_plan_anual_agregar_evidencia(id):
    if not current_user.puede('agregar_evidencias'):
        flash('No tienes permisos para agregar evidencias', 'error')
        return redirect(url_for('sst_plan_anual_actividad_detalle', id=id))
    try:
        titulo = request.form.get('titulo', '').strip()
        descripcion = request.form.get('descripcion', '').strip()
        archivo = request.files.get('archivo')
        if not titulo:
            flash('El título es obligatorio', 'error')
            return redirect(url_for('sst_plan_anual_actividad_detalle', id=id))
        archivo_data = None
        if archivo and archivo.filename != '':
            if allowed_file(archivo.filename):
                archivo_data = guardar_archivo_en_bd(archivo)
            else:
                flash('Tipo de archivo no permitido', 'error')
                return redirect(url_for('sst_plan_anual_actividad_detalle', id=id))
        conn = crear_conexion()
        cursor = conn.cursor()
        if archivo_data:
            cursor.execute("""
                INSERT INTO plan_evidencias (plan_id, titulo, descripcion, archivo_nombre, archivo_tipo, archivo_tamano, archivo_data, usuario_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (id, titulo, descripcion, archivo_data['nombre'], archivo_data['tipo'], archivo_data['tamano'], psycopg2.Binary(archivo_data['data']), current_user.id))
        else:
            cursor.execute("INSERT INTO plan_evidencias (plan_id, titulo, descripcion, usuario_id) VALUES (%s, %s, %s, %s)", (id, titulo, descripcion, current_user.id))
        conn.commit()
        cursor.close()
        conn.close()
        flash('Evidencia agregada exitosamente', 'success')
    except Exception as e:
        flash(f'Error al agregar evidencia: {str(e)}', 'error')
        logger.error(f"Error en sst_plan_anual_agregar_evidencia: {e}")
    return redirect(url_for('sst_plan_anual_actividad_detalle', id=id))

@app.route('/sst/plan-anual/cronograma')
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_plan_anual_cronograma():
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, actividad, ciclo_phva, responsables, estado,
                enero_semana1_p, enero_semana1_e, enero_semana2_p, enero_semana2_e,
                enero_semana3_p, enero_semana3_e, enero_semana4_p, enero_semana4_e,
                febrero_semana1_p, febrero_semana1_e, febrero_semana2_p, febrero_semana2_e,
                febrero_semana3_p, febrero_semana3_e, febrero_semana4_p, febrero_semana4_e,
                marzo_semana1_p, marzo_semana1_e, marzo_semana2_p, marzo_semana2_e,
                marzo_semana3_p, marzo_semana3_e, marzo_semana4_p, marzo_semana4_e,
                abril_semana1_p, abril_semana1_e, abril_semana2_p, abril_semana2_e,
                abril_semana3_p, abril_semana3_e, abril_semana4_p, abril_semana4_e,
                mayo_semana1_p, mayo_semana1_e, mayo_semana2_p, mayo_semana2_e,
                mayo_semana3_p, mayo_semana3_e, mayo_semana4_p, mayo_semana4_e,
                junio_semana1_p, junio_semana1_e, junio_semana2_p, junio_semana2_e,
                junio_semana3_p, junio_semana3_e, junio_semana4_p, junio_semana4_e,
                julio_semana1_p, julio_semana1_e, julio_semana2_p, julio_semana2_e,
                julio_semana3_p, julio_semana3_e, julio_semana4_p, julio_semana4_e,
                agosto_semana1_p, agosto_semana1_e, agosto_semana2_p, agosto_semana2_e,
                agosto_semana3_p, agosto_semana3_e, agosto_semana4_p, agosto_semana4_e,
                septiembre_semana1_p, septiembre_semana1_e, septiembre_semana2_p, septiembre_semana2_e,
                septiembre_semana3_p, septiembre_semana3_e, septiembre_semana4_p, septiembre_semana4_e,
                octubre_semana1_p, octubre_semana1_e, octubre_semana2_p, octubre_semana2_e,
                octubre_semana3_p, octubre_semana3_e, octubre_semana4_p, octubre_semana4_e,
                noviembre_semana1_p, noviembre_semana1_e, noviembre_semana2_p, noviembre_semana2_e,
                noviembre_semana3_p, noviembre_semana3_e, noviembre_semana4_p, noviembre_semana4_e,
                diciembre_semana1_p, diciembre_semana1_e, diciembre_semana2_p, diciembre_semana2_e,
                diciembre_semana3_p, diciembre_semana3_e, diciembre_semana4_p, diciembre_semana4_e
            FROM plan_anual_trabajo ORDER BY ciclo_phva, actividad LIMIT 50
        """)
        actividades = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('sst/plan_anual_cronograma.html', actividades=actividades)
    except Exception as e:
        flash(f'Error al cargar cronograma: {str(e)}', 'error')
        logger.error(f"Error en sst_plan_anual_cronograma: {e}")
        return redirect(url_for('sst_plan_anual'))

@app.route('/sst/plan-anual/inicializar-datos-simple')
@login_required
def sst_inicializar_datos_simple():
    if current_user.rol != 'admin':
        flash('No tienes permisos', 'error')
        return redirect(url_for('sst_dashboard'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM plan_anual_trabajo")
        count = cursor.fetchone()[0]
        if count > 0:
            flash(f'⚠️ Ya existen {count} actividades. No se insertarán duplicados.', 'warning')
            cursor.close()
            conn.close()
            return redirect(url_for('sst_plan_anual'))
        actividades = [
            {'actividad': 'Responsable del Sistema de Gestión de Seguridad y Salud en el Trabajo SG-SST', 'evidencia': 'Documento en el que consta la asignación', 'ciclo_phva': 'Planear', 'articulos': '2.2.4.6.8', 'nivel_pesv': 'N/A', 'responsables': 'SST - COPASST - GERENCIA', 'recursos': 'Tecnologicos, Infraestructura, Humanos', 'estado': 'completado', 'enero_s1_p': True, 'enero_s1_e': True},
            {'actividad': 'Lider del diseño e implementacion del PESV', 'evidencia': 'Documento de asignación del líder', 'ciclo_phva': 'Planear', 'articulos': 'N/A', 'nivel_pesv': 'Paso 1', 'responsables': 'SST - GERENCIA', 'recursos': 'Humanos, Financieros', 'estado': 'en_proceso', 'enero_s1_p': True, 'enero_s1_e': True, 'enero_s2_p': True, 'enero_s2_e': False},
            {'actividad': 'Politica de SST y PESV', 'evidencia': 'Política firmada y comunicada', 'ciclo_phva': 'Planear', 'articulos': '2.2.4.6.5, 2.2.4.6.6', 'nivel_pesv': 'Paso 3', 'responsables': 'SST - COPASST', 'recursos': 'Humanos', 'estado': 'pendiente', 'enero_s1_p': True, 'enero_s2_p': True, 'enero_s3_p': True},
            {'actividad': 'Reuniones mensuales COPASST', 'evidencia': 'Actas de reunión', 'ciclo_phva': 'Hacer', 'articulos': '2.2.4.6.12', 'nivel_pesv': 'N/A', 'responsables': 'SST - COPASST', 'recursos': 'Humanos', 'estado': 'en_proceso', 'enero_s1_p': True, 'febrero_s1_p': True, 'marzo_s1_p': True, 'abril_s1_p': True, 'mayo_s1_p': True, 'junio_s1_p': True},
            {'actividad': 'Revisión trimestral del PESV', 'evidencia': 'Actas de revisión', 'ciclo_phva': 'Verificar', 'articulos': 'N/A', 'nivel_pesv': 'Paso 2', 'responsables': 'Comité de seguridad vial', 'recursos': 'Humanos', 'estado': 'pendiente', 'marzo_s4_p': True, 'junio_s4_p': True, 'septiembre_s4_p': True, 'diciembre_s4_p': True},
            {'actividad': 'Auditoria interna al PESV', 'evidencia': 'Informe de auditoría', 'ciclo_phva': 'Verificar', 'articulos': 'N/A', 'nivel_pesv': 'Paso 22', 'responsables': 'Líder PESV', 'recursos': 'Humanos, Financieros', 'estado': 'pendiente', 'noviembre_s1_p': True, 'noviembre_s2_p': True, 'noviembre_s3_p': True, 'noviembre_s4_p': True},
            {'actividad': 'Acciones preventivas y correctivas', 'evidencia': 'Plan de acción de mejora', 'ciclo_phva': 'Actuar', 'articulos': 'N/A', 'nivel_pesv': 'Paso 23', 'responsables': 'Líder PESV', 'recursos': 'Todos', 'estado': 'pendiente', 'febrero_s1_p': True, 'abril_s1_p': True, 'julio_s1_p': True, 'octubre_s1_p': True},
        ]
        meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
        for act in actividades:
            columnas = ['actividad', 'evidencia', 'ciclo_phva', 'articulos_decreto', 'nivel_pesv', 'responsables', 'recursos', 'estado']
            valores = [act['actividad'], act['evidencia'], act['ciclo_phva'], act['articulos'], act['nivel_pesv'], act['responsables'], act['recursos'], act['estado']]
            for mes in meses:
                for semana in [1, 2, 3, 4]:
                    columnas.append(f'{mes}_semana{semana}_p')
                    valores.append(act.get(f'{mes}_s{semana}_p', False))
                    columnas.append(f'{mes}_semana{semana}_e')
                    valores.append(act.get(f'{mes}_s{semana}_e', False))
            cursor.execute(f"INSERT INTO plan_anual_trabajo ({', '.join(columnas)}) VALUES ({', '.join(['%s'] * len(valores))})", valores)
        conn.commit()
        cursor.close()
        conn.close()
        flash(f'✅ {len(actividades)} actividades del plan anual insertadas correctamente', 'success')
    except Exception as e:
        flash(f'❌ Error al inicializar: {str(e)}', 'error')
        logger.error(f"❌ Error en sst_inicializar_datos_simple: {e}")
    return redirect(url_for('sst_plan_anual'))

@app.route('/sst/plan-anual/verificar-tablas')
@login_required
def sst_verificar_tablas():
    if current_user.rol != 'admin':
        flash('No tienes permisos', 'error')
        return redirect(url_for('sst_dashboard'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name = 'plan_anual_trabajo')")
        tabla_existe = cursor.fetchone()[0]
        if tabla_existe:
            cursor.execute("SELECT COUNT(*) FROM plan_anual_trabajo")
            count = cursor.fetchone()[0]
            flash(f'✅ Tabla plan_anual_trabajo existe con {count} registros', 'success')
        else:
            flash('❌ Tabla plan_anual_trabajo NO existe. Crear tablas primero.', 'error')
        cursor.close()
        conn.close()
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
        logger.error(f"Error en verificar_tablas: {e}")
    return redirect(url_for('sst_plan_anual'))

@app.route('/sst/plan-anual/subir-excel', methods=['GET', 'POST'])
@login_required
def sst_subir_excel():
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos', 'error')
        return redirect_a_modulo_principal()
    if request.method == 'POST':
        try:
            file = request.files.get('excel_file')
            if not file or file.filename == '':
                flash('❌ No se seleccionó ningún archivo', 'error')
                return redirect(url_for('sst_subir_excel'))
            if not file.filename.endswith(('.xlsx', '.xls')):
                flash('❌ El archivo debe ser un Excel (.xlsx o .xls)', 'error')
                return redirect(url_for('sst_subir_excel'))
            excel_path = os.path.join('/tmp', 'Plan_Anual_de_Trabajo_2026.xlsx')
            file.save(excel_path)
            flash('✅ Archivo Excel subido correctamente. Ahora puedes importar los datos.', 'success')
            return redirect(url_for('sst_importar_desde_excel'))
        except Exception as e:
            flash(f'❌ Error al subir archivo: {str(e)}', 'error')
            logger.error(f"Error en sst_subir_excel: {e}")
    return render_template('sst/subir_excel.html')

@app.route('/sst/plan-anual/importar-desde-excel')
@login_required
def sst_importar_desde_excel():
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    try:
        import openpyxl
        excel_path = '/tmp/Plan_Anual_de_Trabajo_2026.xlsx'
        if not os.path.exists(excel_path):
            flash('❌ Archivo Excel no encontrado. Debes subirlo primero.', 'error')
            return redirect(url_for('sst_subir_excel'))
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM plan_anual_trabajo")
        count = cursor.fetchone()[0]
        if count > 10:
            flash(f'⚠️ Ya existen {count} actividades. Elimínalas primero desde la ruta /sst/plan-anual/limpiar-datos', 'warning')
            cursor.close()
            conn.close()
            return redirect(url_for('sst_plan_anual'))
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        actividades = []
        fila_actual = 13
        meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
        while fila_actual <= ws.max_row:
            actividad = ws.cell(fila_actual, 2).value
            evidencia = ws.cell(fila_actual, 3).value
            ciclo_phva = ws.cell(fila_actual, 4).value
            articulos = ws.cell(fila_actual, 5).value
            nivel_pesv = ws.cell(fila_actual, 6).value
            responsables = ws.cell(fila_actual, 7).value
            recursos = ws.cell(fila_actual, 8).value
            if not actividad or isinstance(actividad, str) and any(kw in actividad.upper() for kw in ['PLANEAR', 'HACER', 'VERIFICAR', 'ACTUAR', 'DISEÑO']):
                fila_actual += 1
                continue
            programacion = {}
            col_inicio = 9
            for idx, mes in enumerate(meses):
                programacion[mes] = []
                mes_col_inicio = col_inicio + (idx * 8)
                for semana in range(4):
                    col_p = mes_col_inicio + (semana * 2)
                    col_e = col_p + 1
                    val_p = ws.cell(fila_actual, col_p).value
                    val_e = ws.cell(fila_actual, col_e).value
                    programacion[mes].append({'planificado': val_p in ['x', 'X', True, 1, '1'] if val_p else False, 'ejecutado': val_e in ['x', 'X', True, 1, '1'] if val_e else False})
            actividades.append({'actividad': str(actividad).strip() if actividad else '', 'evidencia': str(evidencia).strip() if evidencia else '', 'ciclo_phva': str(ciclo_phva).strip() if ciclo_phva else '', 'articulos': str(articulos).strip() if articulos else '', 'nivel_pesv': str(nivel_pesv).strip() if nivel_pesv else '', 'responsables': str(responsables).strip() if responsables else '', 'recursos': str(recursos).strip() if recursos else '', 'programacion': programacion})
            fila_actual += 1
        insertadas = 0
        for act in actividades:
            try:
                columnas = ['actividad', 'evidencia', 'ciclo_phva', 'articulos_decreto', 'nivel_pesv', 'responsables', 'recursos', 'estado']
                valores = [act['actividad'][:500] if act['actividad'] else None, act['evidencia'][:500] if act['evidencia'] else None, act['ciclo_phva'][:50] if act['ciclo_phva'] else None, act['articulos'][:200] if act['articulos'] else None, act['nivel_pesv'][:100] if act['nivel_pesv'] else None, act['responsables'][:200] if act['responsables'] else None, act['recursos'][:200] if act['recursos'] else None, 'pendiente']
                for mes in meses:
                    if mes in act['programacion']:
                        semanas = act['programacion'][mes]
                        for semana_idx, semana in enumerate(semanas, 1):
                            columnas.append(f'{mes}_semana{semana_idx}_p')
                            valores.append(semana['planificado'])
                            columnas.append(f'{mes}_semana{semana_idx}_e')
                            valores.append(semana['ejecutado'])
                    else:
                        for semana in range(1, 5):
                            columnas.append(f'{mes}_semana{semana}_p')
                            valores.append(False)
                            columnas.append(f'{mes}_semana{semana}_e')
                            valores.append(False)
                cursor.execute(f"INSERT INTO plan_anual_trabajo ({', '.join(columnas)}) VALUES ({', '.join(['%s'] * len(valores))})", valores)
                insertadas += 1
            except Exception as e:
                logger.error(f"Error insertando: {e}")
        conn.commit()
        cursor.close()
        conn.close()
        flash(f'✅ {insertadas} actividades importadas correctamente desde el Excel', 'success')
    except Exception as e:
        flash(f'❌ Error al importar: {str(e)}', 'error')
        logger.error(f"❌ Error en importar_desde_excel: {e}")
        import traceback
        traceback.print_exc()
    return redirect(url_for('sst_plan_anual'))

@app.route('/sst/plan-anual/limpiar-datos', methods=['GET', 'POST'])
@login_required
def sst_plan_anual_limpiar_datos():
    if current_user.rol != 'admin':
        flash('❌ Solo el administrador puede eliminar todos los datos', 'error')
        return redirect(url_for('sst_plan_anual'))
    if request.method == 'POST':
        confirmacion = request.form.get('confirmacion', '')
        if confirmacion != 'ELIMINAR TODO':
            flash('❌ Debes escribir "ELIMINAR TODO" para confirmar', 'error')
            return redirect(url_for('sst_plan_anual_limpiar_datos'))
        try:
            conn = crear_conexion()
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM plan_anual_trabajo")
            total = cursor.fetchone()[0]
            cursor.execute("DELETE FROM plan_anual_trabajo")
            conn.commit()
            cursor.close()
            conn.close()
            flash(f'✅ {total} actividades eliminadas correctamente', 'success')
            logger.info(f"Admin {current_user.id} eliminó {total} actividades del plan anual")
        except Exception as e:
            flash(f'❌ Error al eliminar: {str(e)}', 'error')
            logger.error(f"Error en sst_plan_anual_limpiar_datos: {e}")
        return redirect(url_for('sst_plan_anual'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) as total, COUNT(*) FILTER (WHERE ciclo_phva = 'Planear') as planear, COUNT(*) FILTER (WHERE ciclo_phva = 'Hacer') as hacer, COUNT(*) FILTER (WHERE ciclo_phva = 'Verificar') as verificar, COUNT(*) FILTER (WHERE ciclo_phva = 'Actuar') as actuar FROM plan_anual_trabajo")
        stats = cursor.fetchone()
        cursor.close()
        conn.close()
        return render_template('sst/plan_anual_limpiar_confirmacion.html', stats=stats)
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
        return redirect(url_for('sst_plan_anual'))

@app.route('/sst/plan-anual/gestionar')
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_plan_anual_gestionar():
    if not current_user.puede('gestionar_plan_anual'):
        flash('No tienes permisos para gestionar actividades', 'error')
        return redirect(url_for('sst_plan_anual'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COUNT(*) as total,
                COUNT(*) FILTER (WHERE ciclo_phva = 'Planear') as planear,
                COUNT(*) FILTER (WHERE ciclo_phva = 'Hacer') as hacer,
                COUNT(*) FILTER (WHERE ciclo_phva = 'Verificar') as verificar,
                COUNT(*) FILTER (WHERE ciclo_phva = 'Actuar') as actuar,
                COUNT(*) FILTER (WHERE estado = 'completado') as completadas,
                COUNT(*) FILTER (WHERE estado = 'en_proceso') as en_proceso,
                COUNT(*) FILTER (WHERE estado = 'pendiente') as pendientes
            FROM plan_anual_trabajo
        """)
        stats = cursor.fetchone()
        filtro_ciclo = request.args.get('ciclo', '')
        filtro_estado = request.args.get('estado', '')
        busqueda = request.args.get('q', '')
        query = "SELECT id, actividad, ciclo_phva, responsables, estado, porcentaje_avance, fecha_actualizacion FROM plan_anual_trabajo WHERE 1=1"
        params = []
        if filtro_ciclo:
            query += " AND ciclo_phva = %s"
            params.append(filtro_ciclo)
        if filtro_estado:
            query += " AND estado = %s"
            params.append(filtro_estado)
        if busqueda:
            query += " AND (actividad ILIKE %s OR responsables ILIKE %s)"
            params.extend([f'%{busqueda}%', f'%{busqueda}%'])
        query += " ORDER BY ciclo_phva, actividad LIMIT 100"
        cursor.execute(query, params)
        actividades = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('sst/plan_anual_gestionar.html', stats=stats, actividades=actividades, filtro_ciclo=filtro_ciclo, filtro_estado=filtro_estado, busqueda=busqueda)
    except Exception as e:
        flash(f'❌ Error al cargar panel de gestión: {str(e)}', 'error')
        logger.error(f"Error en sst_plan_anual_gestionar: {e}")
        return redirect(url_for('sst_plan_anual'))

@app.route('/sst/plan-anual/actividad/nueva', methods=['GET', 'POST'])
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_plan_anual_nueva_actividad():
    if not current_user.puede('gestionar_plan_anual'):
        flash('No tienes permisos para crear actividades', 'error')
        return redirect(url_for('sst_plan_anual_actividades'))
    try:
        if request.method == 'POST':
            actividad = request.form.get('actividad', '').strip()
            evidencia = request.form.get('evidencia', '').strip()
            ciclo_phva = request.form.get('ciclo_phva', '').strip()
            articulos = request.form.get('articulos_decreto', '').strip()
            nivel_pesv = request.form.get('nivel_pesv', '').strip()
            responsables = request.form.get('responsables', '').strip()
            recursos = request.form.get('recursos', '').strip()
            observaciones = request.form.get('observaciones', '').strip()
            estado = request.form.get('estado', 'pendiente')
            if not actividad or not ciclo_phva:
                flash('❌ La actividad y el ciclo PHVA son obligatorios', 'error')
                return render_template('sst/plan_anual_nueva.html')
            conn = crear_conexion()
            cursor = conn.cursor()
            columnas = ['actividad', 'evidencia', 'ciclo_phva', 'articulos_decreto', 'nivel_pesv', 'responsables', 'recursos', 'observaciones', 'estado', 'usuario_actualizacion']
            valores = [actividad[:500], evidencia[:500] if evidencia else None, ciclo_phva[:50], articulos[:200] if articulos else None, nivel_pesv[:100] if nivel_pesv else None, responsables[:200] if responsables else None, recursos[:200] if recursos else None, observaciones, estado, current_user.id]
            meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
            for mes in meses:
                for semana in range(1, 5):
                    columnas.append(f'{mes}_semana{semana}_p')
                    valores.append(request.form.get(f'{mes}_semana{semana}_p') == 'on')
                    columnas.append(f'{mes}_semana{semana}_e')
                    valores.append(request.form.get(f'{mes}_semana{semana}_e') == 'on')
            cursor.execute(f"INSERT INTO plan_anual_trabajo ({', '.join(columnas)}) VALUES ({', '.join(['%s'] * len(valores))}) RETURNING id", valores)
            new_id = cursor.fetchone()[0]
            conn.commit()
            actualizar_porcentaje_avance(new_id)
            cursor.close()
            conn.close()
            flash('✅ Actividad creada correctamente', 'success')
            return redirect(url_for('sst_plan_anual_actividad_detalle', id=new_id))
        return render_template('sst/plan_anual_nueva.html')
    except Exception as e:
        flash(f'❌ Error al crear actividad: {str(e)}', 'error')
        logger.error(f"Error en sst_plan_anual_nueva_actividad: {e}")
        return redirect(url_for('sst_plan_anual_actividades'))

@app.route('/sst/plan-anual/actividad/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_plan_anual_editar_actividad(id):
    if not current_user.puede('gestionar_plan_anual'):
        flash('No tienes permisos para editar actividades', 'error')
        return redirect(url_for('sst_plan_anual_actividades'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        if request.method == 'POST':
            actividad = request.form.get('actividad', '').strip()
            evidencia = request.form.get('evidencia', '').strip()
            ciclo_phva = request.form.get('ciclo_phva', '').strip()
            articulos = request.form.get('articulos_decreto', '').strip()
            nivel_pesv = request.form.get('nivel_pesv', '').strip()
            responsables = request.form.get('responsables', '').strip()
            recursos = request.form.get('recursos', '').strip()
            observaciones = request.form.get('observaciones', '').strip()
            estado = request.form.get('estado', 'pendiente')
            if not actividad or not ciclo_phva:
                flash('❌ La actividad y el ciclo PHVA son obligatorios', 'error')
                return redirect(url_for('sst_plan_anual_editar_actividad', id=id))
            cursor.execute("UPDATE plan_anual_trabajo SET actividad = %s, evidencia = %s, ciclo_phva = %s, articulos_decreto = %s, nivel_pesv = %s, responsables = %s, recursos = %s, observaciones = %s, estado = %s, fecha_actualizacion = CURRENT_TIMESTAMP, usuario_actualizacion = %s WHERE id = %s",
                          (actividad[:500], evidencia[:500] if evidencia else None, ciclo_phva[:50], articulos[:200] if articulos else None, nivel_pesv[:100] if nivel_pesv else None, responsables[:200] if responsables else None, recursos[:200] if recursos else None, observaciones, estado, current_user.id, id))
            meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
            for mes in meses:
                for semana in range(1, 5):
                    key_p = f'{mes}_semana{semana}_p'
                    key_e = f'{mes}_semana{semana}_e'
                    cursor.execute(f"UPDATE plan_anual_trabajo SET {key_p} = %s, {key_e} = %s WHERE id = %s", (request.form.get(key_p) == 'on', request.form.get(key_e) == 'on', id))
            conn.commit()
            actualizar_porcentaje_avance(id)
            flash('✅ Actividad actualizada correctamente', 'success')
            cursor.close()
            conn.close()
            return redirect(url_for('sst_plan_anual_actividad_detalle', id=id))
        cursor.execute("SELECT * FROM plan_anual_trabajo WHERE id = %s", (id,))
        actividad_data = cursor.fetchone()
        if not actividad_data:
            flash('❌ Actividad no encontrada', 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('sst_plan_anual_actividades'))
        actividad = {
            'id': actividad_data[0], 'actividad': actividad_data[1], 'evidencia': actividad_data[2],
            'ciclo_phva': actividad_data[3], 'articulos_decreto': actividad_data[4],
            'nivel_pesv': actividad_data[5], 'responsables': actividad_data[6], 'recursos': actividad_data[7],
            'observaciones': actividad_data[103] if len(actividad_data) > 103 else '',
            'estado': actividad_data[104] if len(actividad_data) > 104 else 'pendiente',
            'porcentaje_avance': actividad_data[105] if len(actividad_data) > 105 else 0
        }
        meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
        programacion = {}
        col_offset = 8
        for i, mes in enumerate(meses):
            programacion[mes] = []
            for semana in range(1, 5):
                idx_p = col_offset + (i * 8) + ((semana - 1) * 2)
                idx_e = idx_p + 1
                programacion[mes].append({'semana': semana, 'planificado': actividad_data[idx_p] if idx_p < len(actividad_data) else False, 'ejecutado': actividad_data[idx_e] if idx_e < len(actividad_data) else False})
        actividad['programacion'] = programacion
        cursor.close()
        conn.close()
        return render_template('sst/plan_anual_editar.html', actividad=actividad)
    except Exception as e:
        flash(f'❌ Error al editar actividad: {str(e)}', 'error')
        logger.error(f"Error en sst_plan_anual_editar_actividad: {e}")
        return redirect(url_for('sst_plan_anual_actividades'))

@app.route('/sst/plan-anual/actividad/<int:id>/eliminar', methods=['POST'])
@login_required
def sst_plan_anual_eliminar_actividad(id):
    if not current_user.puede('gestionar_plan_anual'):
        flash('No tienes permisos para eliminar actividades', 'error')
        return redirect(url_for('sst_plan_anual_actividades'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("SELECT actividad FROM plan_anual_trabajo WHERE id = %s", (id,))
        actividad = cursor.fetchone()
        if not actividad:
            flash('❌ Actividad no encontrada', 'error')
        else:
            cursor.execute("DELETE FROM plan_anual_trabajo WHERE id = %s", (id,))
            conn.commit()
            flash(f'✅ Actividad "{actividad[0][:50]}..." eliminada correctamente', 'success')
        cursor.close()
        conn.close()
    except Exception as e:
        flash(f'❌ Error al eliminar: {str(e)}', 'error')
        logger.error(f"Error en sst_plan_anual_eliminar_actividad: {e}")
    return redirect(url_for('sst_plan_anual_actividades'))

@app.route('/sst/plan-anual/gestionar/limpiar-masivo', methods=['POST'])
@login_required
def sst_plan_anual_limpiar_masivo():
    if current_user.rol != 'admin':
        flash('Solo el administrador puede eliminar masivamente', 'error')
        return redirect(url_for('sst_plan_anual_gestionar'))
    try:
        ids = request.form.getlist('actividad_ids')
        if not ids:
            flash('❌ No se seleccionaron actividades', 'warning')
            return redirect(url_for('sst_plan_anual_gestionar'))
        conn = crear_conexion()
        cursor = conn.cursor()
        ids_int = [int(id) for id in ids]
        placeholders = ', '.join(['%s'] * len(ids_int))
        cursor.execute(f"DELETE FROM plan_anual_trabajo WHERE id IN ({placeholders})", ids_int)
        eliminadas = cursor.rowcount
        conn.commit()
        cursor.close()
        conn.close()
        flash(f'✅ {eliminadas} actividades eliminadas correctamente', 'success')
    except Exception as e:
        flash(f'❌ Error en eliminación masiva: {str(e)}', 'error')
        logger.error(f"Error en sst_plan_anual_limpiar_masivo: {e}")
    return redirect(url_for('sst_plan_anual_gestionar'))

@app.route('/sst/plan-anual/evidencia/<int:id>/descargar')
@login_required
@retry_on_ssl_error(max_retries=2, delay=2)
def sst_descargar_evidencia(id):
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("SELECT archivo_nombre, archivo_tipo, archivo_data FROM plan_evidencias WHERE id = %s", (id,))
        evidencia = cursor.fetchone()
        cursor.close()
        conn.close()
        if not evidencia or not evidencia[2]:
            flash('Archivo de evidencia no encontrado', 'error')
            return redirect(url_for('sst_plan_anual_actividades'))
        file_data = BytesIO(bytes(evidencia[2]))
        return send_file(file_data, mimetype=evidencia[1] or 'application/octet-stream', as_attachment=True, download_name=evidencia[0])
    except Exception as e:
        flash(f'Error al descargar evidencia: {str(e)}', 'error')
        logger.error(f"Error en sst_descargar_evidencia: {e}")
        return redirect(url_for('sst_plan_anual_actividades'))

@app.route('/sst/plan-anual/evidencia/<int:id>/ver')
@login_required
def sst_ver_evidencia(id):
    if not current_user.puede('acceder_sst'):
        return redirect_a_modulo_principal()
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("SELECT archivo_nombre, archivo_tipo, archivo_data FROM plan_evidencias WHERE id = %s", (id,))
        evidencia = cursor.fetchone()
        cursor.close()
        conn.close()
        if not evidencia or not evidencia[2]:
            flash('Archivo no encontrado', 'error')
            return redirect(url_for('sst_plan_anual_actividades'))
        file_data = BytesIO(bytes(evidencia[2]))
        return send_file(file_data, mimetype=evidencia[1] or 'application/octet-stream', as_attachment=False, download_name=evidencia[0])
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('sst_plan_anual_actividades'))

@app.route('/sst/plan-anual/actividad/<int:id>/seguimiento', methods=['POST'])
@login_required
def sst_plan_anual_agregar_seguimiento(id):
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    try:
        comentario = request.form.get('comentario', '').strip()
        tipo = request.form.get('tipo', 'comentario').strip()
        if not comentario:
            flash('❌ El comentario no puede estar vacío', 'error')
            return redirect(url_for('sst_plan_anual_actividad_detalle', id=id))
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO plan_seguimiento (actividad_id, comentario, tipo, usuario_id, fecha) VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP)", (id, comentario, tipo, current_user.id))
        conn.commit()
        cursor.close()
        conn.close()
        flash('✅ Seguimiento agregado correctamente', 'success')
    except Exception as e:
        flash(f'❌ Error al agregar seguimiento: {str(e)}', 'error')
        logger.error(f"Error en sst_plan_anual_agregar_seguimiento: {e}")
    return redirect(url_for('sst_plan_anual_actividad_detalle', id=id))

@app.route('/sst/plan-anual/evidencia/<int:id>/eliminar', methods=['POST'])
@login_required
def sst_eliminar_evidencia(id):
    if not current_user.puede('acceder_sst'):
        flash('No tienes permisos para acceder al módulo de SST', 'error')
        return redirect_a_modulo_principal()
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("SELECT plan_id FROM plan_evidencias WHERE id = %s", (id,))
        resultado = cursor.fetchone()
        if not resultado:
            flash('❌ Evidencia no encontrada', 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('sst_plan_anual_actividades'))
        plan_id = resultado[0]
        cursor.execute("DELETE FROM plan_evidencias WHERE id = %s", (id,))
        conn.commit()
        cursor.close()
        conn.close()
        flash('✅ Evidencia eliminada correctamente', 'success')
    except Exception as e:
        flash(f'❌ Error al eliminar evidencia: {str(e)}', 'error')
        logger.error(f"Error en sst_eliminar_evidencia: {e}")
        return redirect(url_for('sst_plan_anual_actividades'))
    return redirect(url_for('sst_plan_anual_actividad_detalle', id=plan_id))


# ==========================================
# MÓDULO DE RECURSOS HUMANOS (RH)
# ==========================================

@app.route('/rh_dashboard')
@login_required
def rh_dashboard_redirect():
    """Redirección a /rh para compatibilidad"""
    return redirect(url_for('rh_dashboard'))

@app.route('/rh/empleados')
@login_required
def rh_empleados():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        busqueda = request.args.get('q', '')
        departamento_filtro = request.args.get('departamento', '')
        estado_filtro = request.args.get('estado', '')
        
        # Consulta CORREGIDA con JOINs
        query = """
            SELECT 
                e.id, 
                e.tipo_documento, 
                e.documento, 
                e.primer_nombre, 
                e.segundo_nombre,
                e.primer_apellido, 
                e.segundo_apellido, 
                e.email, 
                e.telefono,
                e.celular,
                e.direccion,
                e.fecha_nacimiento,
                COALESCE(c.nombre, 'Sin asignar') as cargo,
                COALESCE(d.nombre, 'Sin asignar') as departamento,
                e.fecha_ingreso, 
                e.fecha_retiro,
                e.estado, 
                e.salario
            FROM rh_empleados e
            LEFT JOIN rh_cargos c ON e.cargo_id = c.id
            LEFT JOIN rh_departamentos d ON e.departamento_id = d.id
            WHERE 1=1
        """
        params = []
        
        if busqueda:
            query += " AND (e.primer_nombre ILIKE %s OR e.primer_apellido ILIKE %s OR e.documento ILIKE %s)"
            params.extend([f'%{busqueda}%', f'%{busqueda}%', f'%{busqueda}%'])
        
        if departamento_filtro:
            query += " AND d.nombre ILIKE %s"
            params.append(f'%{departamento_filtro}%')
        
        if estado_filtro:
            query += " AND e.estado = %s"
            params.append(estado_filtro)
        
        query += " ORDER BY e.primer_apellido, e.primer_nombre"
        
        cursor.execute(query, params)
        empleados = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return render_template('rh/rh_empleados.html', empleados=empleados)
    except Exception as e:
        logger.error(f"Error en rh_empleados: {e}")
        flash('Error al cargar los empleados', 'error')
        return render_template('rh/rh_empleados.html', empleados=[])

@app.route('/rh/empleado/<int:id>')
@login_required
def rh_empleado_detalle(id):
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        
        # Consulta CORREGIDA - usando SOLO las columnas que existen en tu BD
        cursor.execute("""
            SELECT 
                e.id, 
                e.tipo_documento, 
                e.documento, 
                e.primer_nombre, 
                e.segundo_nombre,
                e.primer_apellido, 
                e.segundo_apellido, 
                e.email, 
                e.telefono, 
                e.celular,
                e.direccion, 
                e.fecha_nacimiento, 
                COALESCE(c.nombre, 'No asignado') as cargo,
                COALESCE(d.nombre, 'No asignado') as departamento,
                e.fecha_ingreso, 
                e.fecha_retiro, 
                e.estado, 
                e.salario
            FROM rh_empleados e
            LEFT JOIN rh_cargos c ON e.cargo_id = c.id
            LEFT JOIN rh_departamentos d ON e.departamento_id = d.id
            WHERE e.id = %s
        """, (id,))
        
        empleado = cursor.fetchone()
        
        if not empleado:
            flash('Empleado no encontrado', 'error')
            return redirect(url_for('rh_empleados'))
        
        # Obtener contratos del empleado
        cursor.execute("""
            SELECT 
                id, 
                tipo_contrato, 
                fecha_inicio, 
                fecha_fin, 
                salario_contratado, 
                estado,
                archivo_pdf
            FROM rh_contratos 
            WHERE empleado_id = %s 
            ORDER BY fecha_inicio DESC
        """, (id,))
        contratos = cursor.fetchall()
        
        # Obtener capacitaciones del empleado
        try:
            cursor.execute("""
                SELECT 
                    c.id, 
                    c.titulo, 
                    c.fecha_inicio, 
                    c.fecha_fin, 
                    cp.asistio,
                    cp.nota,
                    cp.certificado_generado
                FROM rh_capacitacion_participantes cp
                JOIN rh_capacitaciones c ON cp.capacitacion_id = c.id
                WHERE cp.empleado_id = %s 
                ORDER BY c.fecha_inicio DESC
                LIMIT 5
            """, (id,))
            capacitaciones = cursor.fetchall()
        except Exception as e:
            logger.warning(f"No se pudieron cargar capacitaciones: {e}")
            capacitaciones = []
        
        # Obtener evaluaciones del empleado
        try:
            cursor.execute("""
                SELECT 
                    id, 
                    periodo, 
                    fecha_evaluacion, 
                    puntaje_total, 
                    estado
                FROM rh_evaluaciones 
                WHERE empleado_id = %s 
                ORDER BY fecha_evaluacion DESC
                LIMIT 5
            """, (id,))
            evaluaciones = cursor.fetchall()
        except Exception as e:
            logger.warning(f"No se pudieron cargar evaluaciones: {e}")
            evaluaciones = []
        
        cursor.close()
        conn.close()
        
        return render_template('rh/rh_empleado_detalle.html',
                             empleado=empleado, 
                             contratos=contratos, 
                             capacitaciones=capacitaciones,
                             evaluaciones=evaluaciones)
                             
    except Exception as e:
        logger.error(f"Error en rh_empleado_detalle: {e}")
        flash('Error al cargar el detalle del empleado', 'error')
        return redirect(url_for('rh_empleados'))

@app.route('/rh/empleado/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def rh_empleado_editar(id):
    if not current_user.puede('gestionar_rh'):
        flash('No tienes permisos para editar empleados', 'error')
        return redirect(url_for('rh_empleados'))
    
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        
        # Obtener cargos y departamentos para los selects
        cursor.execute("SELECT id, nombre FROM rh_cargos WHERE activo = TRUE ORDER BY nombre")
        cargos = cursor.fetchall()
        
        cursor.execute("SELECT id, nombre FROM rh_departamentos WHERE activo = TRUE ORDER BY nombre")
        departamentos = cursor.fetchall()
        
        if request.method == 'POST':
            # Obtener datos del formulario - SOLO columnas que existen
            tipo_documento = request.form.get('tipo_documento', 'CC')
            documento = request.form.get('documento', '').strip()
            primer_nombre = request.form.get('primer_nombre', '').strip()
            segundo_nombre = request.form.get('segundo_nombre', '').strip() or None
            primer_apellido = request.form.get('primer_apellido', '').strip()
            segundo_apellido = request.form.get('segundo_apellido', '').strip() or None
            email = request.form.get('email', '').strip() or None
            telefono = request.form.get('telefono', '').strip() or None
            celular = request.form.get('celular', '').strip() or None
            direccion = request.form.get('direccion', '').strip() or None
            fecha_nacimiento = request.form.get('fecha_nacimiento', '').strip() or None
            cargo_id = request.form.get('cargo_id', '').strip()
            departamento_id = request.form.get('departamento_id', '').strip()
            fecha_ingreso = request.form.get('fecha_ingreso', '').strip() or None
            fecha_retiro = request.form.get('fecha_retiro', '').strip() or None
            salario = request.form.get('salario', '').strip() or None
            estado = request.form.get('estado', 'activo')
            
            # Validar campos obligatorios
            if not documento:
                flash('❌ El número de documento es obligatorio', 'error')
                return render_template('rh/rh_empleado_editar.html', 
                                     empleado=None, cargos=cargos, departamentos=departamentos)
            
            if not primer_nombre:
                flash('❌ El primer nombre es obligatorio', 'error')
                return render_template('rh/rh_empleado_editar.html', 
                                     empleado=None, cargos=cargos, departamentos=departamentos)
            
            if not primer_apellido:
                flash('❌ El primer apellido es obligatorio', 'error')
                return render_template('rh/rh_empleado_editar.html', 
                                     empleado=None, cargos=cargos, departamentos=departamentos)
            
            # Validar que el documento no exista en otro empleado
            cursor.execute("SELECT id FROM rh_empleados WHERE documento = %s AND id != %s", (documento, id))
            if cursor.fetchone():
                flash(f'❌ Ya existe otro empleado con el documento {documento}', 'error')
                return render_template('rh/rh_empleado_editar.html', 
                                     empleado=None, cargos=cargos, departamentos=departamentos)
            
            # Validar email único
            if email:
                cursor.execute("SELECT id FROM rh_empleados WHERE email = %s AND id != %s", (email, id))
                if cursor.fetchone():
                    flash(f'❌ Ya existe otro empleado con el email {email}', 'error')
                    return render_template('rh/rh_empleado_editar.html', 
                                         empleado=None, cargos=cargos, departamentos=departamentos)
            
            # Convertir valores
            cargo_id = int(cargo_id) if cargo_id and cargo_id.isdigit() else None
            departamento_id = int(departamento_id) if departamento_id and departamento_id.isdigit() else None
            salario = float(salario) if salario and salario.replace('.', '').isdigit() else None
            
            # Actualizar empleado - SOLO columnas que existen
            cursor.execute("""
                UPDATE rh_empleados SET
                    tipo_documento = %s,
                    documento = %s,
                    primer_nombre = %s,
                    segundo_nombre = %s,
                    primer_apellido = %s,
                    segundo_apellido = %s,
                    email = %s,
                    telefono = %s,
                    celular = %s,
                    direccion = %s,
                    fecha_nacimiento = %s,
                    cargo_id = %s,
                    departamento_id = %s,
                    fecha_ingreso = %s,
                    fecha_retiro = %s,
                    salario = %s,
                    estado = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (
                tipo_documento, documento, primer_nombre, segundo_nombre,
                primer_apellido, segundo_apellido, email, telefono, celular,
                direccion, fecha_nacimiento, cargo_id, departamento_id,
                fecha_ingreso, fecha_retiro, salario, estado, id
            ))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            flash('✅ Empleado actualizado correctamente', 'success')
            return redirect(url_for('rh_empleado_detalle', id=id))
        
        # GET - Obtener datos del empleado
        cursor.execute("""
            SELECT 
                e.id, e.tipo_documento, e.documento, e.primer_nombre, e.segundo_nombre,
                e.primer_apellido, e.segundo_apellido, e.email, e.telefono, e.celular,
                e.direccion, e.fecha_nacimiento, e.cargo_id, e.departamento_id,
                e.fecha_ingreso, e.fecha_retiro, e.estado, e.salario
            FROM rh_empleados e
            WHERE e.id = %s
        """, (id,))
        
        empleado = cursor.fetchone()
        
        if not empleado:
            flash('Empleado no encontrado', 'error')
            return redirect(url_for('rh_empleados'))
        
        cursor.close()
        conn.close()
        
        return render_template('rh/rh_empleado_editar.html', 
                             empleado=empleado, 
                             cargos=cargos, 
                             departamentos=departamentos)
                             
    except Exception as e:
        logger.error(f"Error en rh_empleado_editar: {e}")
        flash('Error al editar el empleado', 'error')
        return redirect(url_for('rh_empleados'))

@app.route('/rh/empleado/nuevo', methods=['GET', 'POST'])
@login_required
def rh_empleado_nuevo():
    if not current_user.puede('gestionar_rh'):
        flash('No tienes permisos para crear empleados', 'error')
        return redirect(url_for('rh_empleados'))
    
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        
        # Obtener cargos y departamentos para los selects
        cursor.execute("SELECT id, nombre FROM rh_cargos WHERE activo = TRUE ORDER BY nombre")
        cargos = cursor.fetchall()
        
        cursor.execute("SELECT id, nombre FROM rh_departamentos WHERE activo = TRUE ORDER BY nombre")
        departamentos = cursor.fetchall()
        
        if request.method == 'POST':
            # Obtener datos del formulario
            tipo_documento = request.form.get('tipo_documento', 'CC')
            documento = request.form.get('documento', '').strip()
            primer_nombre = request.form.get('primer_nombre', '').strip()
            segundo_nombre = request.form.get('segundo_nombre', '').strip() or None
            primer_apellido = request.form.get('primer_apellido', '').strip()
            segundo_apellido = request.form.get('segundo_apellido', '').strip() or None
            email = request.form.get('email', '').strip() or None
            telefono = request.form.get('telefono', '').strip() or None
            celular = request.form.get('celular', '').strip() or None
            direccion = request.form.get('direccion', '').strip() or None
            fecha_nacimiento = request.form.get('fecha_nacimiento', '').strip() or None
            cargo_id = request.form.get('cargo_id', '').strip()
            departamento_id = request.form.get('departamento_id', '').strip()
            fecha_ingreso = request.form.get('fecha_ingreso', '').strip() or None
            salario = request.form.get('salario', '').strip() or None
            # tipo_contrato NO va aquí - se maneja en rh_contratos
            estado = request.form.get('estado', 'activo')
            
            # Validar campos obligatorios
            if not documento:
                flash('❌ El número de documento es obligatorio', 'error')
                return render_template('rh/rh_empleado_nuevo.html', cargos=cargos, departamentos=departamentos)
            
            if not primer_nombre:
                flash('❌ El primer nombre es obligatorio', 'error')
                return render_template('rh/rh_empleado_nuevo.html', cargos=cargos, departamentos=departamentos)
            
            if not primer_apellido:
                flash('❌ El primer apellido es obligatorio', 'error')
                return render_template('rh/rh_empleado_nuevo.html', cargos=cargos, departamentos=departamentos)
            
            # Validar que el documento no exista
            cursor.execute("SELECT id FROM rh_empleados WHERE documento = %s", (documento,))
            if cursor.fetchone():
                flash(f'❌ Ya existe un empleado con el documento {documento}', 'error')
                return render_template('rh/rh_empleado_nuevo.html', cargos=cargos, departamentos=departamentos)
            
            # Validar email único si se proporcionó
            if email:
                cursor.execute("SELECT id FROM rh_empleados WHERE email = %s", (email,))
                if cursor.fetchone():
                    flash(f'❌ Ya existe un empleado con el email {email}', 'error')
                    return render_template('rh/rh_empleado_nuevo.html', cargos=cargos, departamentos=departamentos)
            
            # Convertir valores vacíos a None
            cargo_id = int(cargo_id) if cargo_id and cargo_id.isdigit() else None
            departamento_id = int(departamento_id) if departamento_id and departamento_id.isdigit() else None
            salario = float(salario) if salario and salario.replace('.', '').isdigit() else None
            
            # Insertar empleado - SIN tipo_contrato
            cursor.execute("""
                INSERT INTO rh_empleados (
                    tipo_documento, documento, primer_nombre, segundo_nombre,
                    primer_apellido, segundo_apellido, email, telefono, celular,
                    direccion, fecha_nacimiento, cargo_id, departamento_id,
                    fecha_ingreso, salario, estado, created_at, updated_at
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP
                ) RETURNING id
            """, (
                tipo_documento, documento, primer_nombre, segundo_nombre,
                primer_apellido, segundo_apellido, email, telefono, celular,
                direccion, fecha_nacimiento, cargo_id, departamento_id,
                fecha_ingreso, salario, estado
            ))
            
            new_id = cursor.fetchone()[0]
            conn.commit()
            
            # Si se proporcionó tipo_contrato, crear contrato automáticamente
            tipo_contrato_form = request.form.get('tipo_contrato', '').strip()
            if tipo_contrato_form and fecha_ingreso:
                cursor.execute("""
                    INSERT INTO rh_contratos (empleado_id, tipo_contrato, fecha_inicio, salario_contratado, estado)
                    VALUES (%s, %s, %s, %s, 'activo')
                """, (new_id, tipo_contrato_form, fecha_ingreso, salario))
                conn.commit()
            
            cursor.close()
            conn.close()
            
            flash('✅ Empleado creado correctamente', 'success')
            return redirect(url_for('rh_empleado_detalle', id=new_id))
        
        cursor.close()
        conn.close()
        return render_template('rh/rh_empleado_nuevo.html', cargos=cargos, departamentos=departamentos)
        
    except Exception as e:
        logger.error(f"Error en rh_empleado_nuevo: {e}")
        flash(f'Error al crear el empleado: {str(e)}', 'error')
        return redirect(url_for('rh_empleados'))

@app.route('/rh/contratos')
@login_required
def rh_contratos():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                c.id, 
                e.primer_nombre, 
                e.primer_apellido, 
                e.documento,
                c.tipo_contrato, 
                c.fecha_inicio, 
                c.fecha_fin, 
                c.salario_contratado, 
                c.estado,
                c.archivo_pdf
            FROM rh_contratos c
            JOIN rh_empleados e ON c.empleado_id = e.id
            ORDER BY c.fecha_inicio DESC
        """)
        contratos = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('rh/rh_contratos.html', contratos=contratos)
    except Exception as e:
        logger.error(f"Error en rh_contratos: {e}")
        return render_template('rh/rh_contratos.html', contratos=[])

@app.route('/rh/contrato/<int:id>')
@login_required
def rh_contrato_detalle(id):
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT 
                c.*,
                e.primer_nombre, 
                e.primer_apellido, 
                e.documento, 
                e.cargo_id,
                COALESCE(carg.nombre, 'Sin cargo') as cargo_nombre
            FROM rh_contratos c 
            JOIN rh_empleados e ON c.empleado_id = e.id
            LEFT JOIN rh_cargos carg ON e.cargo_id = carg.id
            WHERE c.id = %s
        """, (id,))
        
        contrato = cursor.fetchone()
        
        if not contrato:
            flash('Contrato no encontrado', 'error')
            return redirect(url_for('rh_contratos'))
        
        cursor.close()
        conn.close()
        
        return render_template('rh/rh_contrato_detalle.html', contrato=contrato)
        
    except Exception as e:
        logger.error(f"Error en rh_contrato_detalle: {e}")
        flash('Error al cargar el detalle del contrato', 'error')
        return redirect(url_for('rh_contratos'))
        
@app.route('/rh/contrato/nuevo', methods=['GET', 'POST'])
@login_required
def rh_contrato_nuevo():
    if not current_user.puede('gestionar_rh'):
        flash('No tienes permisos para crear contratos', 'error')
        return redirect(url_for('rh_contratos'))
    
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        
        # Obtener lista de empleados activos para el select
        cursor.execute("""
            SELECT id, primer_nombre, primer_apellido, documento 
            FROM rh_empleados 
            WHERE estado = 'activo'
            ORDER BY primer_apellido, primer_nombre
        """)
        empleados = cursor.fetchall()
        
        if request.method == 'POST':
            empleado_id = request.form.get('empleado_id')
            tipo_contrato = request.form.get('tipo_contrato')
            fecha_inicio = request.form.get('fecha_inicio')
            fecha_fin = request.form.get('fecha_fin') or None
            salario_contratado = request.form.get('salario_contratado')
            estado = request.form.get('estado', 'activo')
            
            # Validaciones
            if not empleado_id:
                flash('❌ Debe seleccionar un empleado', 'error')
                return render_template('rh/rh_contrato_nuevo.html', empleados=empleados)
            
            if not tipo_contrato:
                flash('❌ Debe seleccionar un tipo de contrato', 'error')
                return render_template('rh/rh_contrato_nuevo.html', empleados=empleados)
            
            if not fecha_inicio:
                flash('❌ La fecha de inicio es obligatoria', 'error')
                return render_template('rh/rh_contrato_nuevo.html', empleados=empleados)
            
            if not salario_contratado:
                flash('❌ El salario contratado es obligatorio', 'error')
                return render_template('rh/rh_contrato_nuevo.html', empleados=empleados)
            
            # Convertir salario a número
            try:
                salario_contratado = float(salario_contratado.replace(',', '.'))
            except:
                salario_contratado = float(salario_contratado) if salario_contratado else None
            
            # Insertar contrato - SOLO columnas que existen en tu BD
            cursor.execute("""
                INSERT INTO rh_contratos (
                    empleado_id, 
                    tipo_contrato, 
                    fecha_inicio, 
                    fecha_fin, 
                    salario_contratado, 
                    estado,
                    created_at
                ) VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP) RETURNING id
            """, (empleado_id, tipo_contrato, fecha_inicio, fecha_fin, salario_contratado, estado))
            
            new_id = cursor.fetchone()[0]
            
            # Actualizar también los datos del empleado (fecha_ingreso, salario, tipo_contrato)
            cursor.execute("""
                UPDATE rh_empleados 
                SET fecha_ingreso = COALESCE(fecha_ingreso, %s),
                    salario = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (fecha_inicio, salario_contratado, empleado_id))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            flash('✅ Contrato creado correctamente', 'success')
            return redirect(url_for('rh_contrato_detalle', id=new_id))
        
        cursor.close()
        conn.close()
        return render_template('rh/rh_contrato_nuevo.html', empleados=empleados)
        
    except Exception as e:
        logger.error(f"Error en rh_contrato_nuevo: {e}")
        flash(f'Error al crear el contrato: {str(e)}', 'error')
        return redirect(url_for('rh_contratos'))
        
@app.route('/rh')
@login_required
def rh_dashboard():
    """Dashboard principal de Recursos Humanos"""
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                (SELECT COUNT(*) FROM rh_empleados WHERE estado = 'activo') as total_empleados,
                (SELECT COUNT(*) FROM rh_empleados WHERE fecha_ingreso >= DATE_TRUNC('month', CURRENT_DATE)) as ingresos_mes,
                (SELECT COUNT(*) FROM rh_empleados WHERE fecha_retiro >= DATE_TRUNC('month', CURRENT_DATE)) as retiros_mes
        """)
        stats = cursor.fetchone()
        cursor.execute("""
            SELECT id, nombre, responsable_id, estado, avance, fecha_limite
            FROM rh_procesos WHERE estado != 'completado' ORDER BY fecha_limite ASC LIMIT 5
        """)
        procesos = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('rh/rh_dashboard.html',
                             total_empleados=stats[0] if stats else 0,
                             ingresos_mes=stats[1] if stats else 0,
                             retiros_mes=stats[2] if stats else 0,
                             procesos_activos=procesos)
    except Exception as e:
        logger.error(f"Error en rh_dashboard: {e}")
        return render_template('rh/rh_dashboard.html',
                             total_empleados=0, ingresos_mes=0, retiros_mes=0, procesos_activos=[])

@app.route('/rh/procesos', methods=['GET', 'POST'])
@login_required
def rh_procesos():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    
    # Procesar POST (crear nuevo proceso)
    if request.method == 'POST':
        if not current_user.puede('gestionar_rh'):
            flash('No tienes permisos para crear procesos', 'error')
            return redirect(url_for('rh_procesos'))
        
        try:
            conn = crear_conexion()
            cursor = conn.cursor()
            
            nombre = request.form.get('nombre')
            descripcion = request.form.get('descripcion')
            tipo = request.form.get('tipo')
            prioridad = request.form.get('prioridad', 'media')
            fecha_limite = request.form.get('fecha_limite') or None
            
            if not nombre or not tipo:
                flash('Nombre y tipo son obligatorios', 'error')
                cursor.close()
                conn.close()
                return redirect(url_for('rh_procesos'))
            
            cursor.execute("""
                INSERT INTO rh_procesos (nombre, descripcion, tipo, prioridad, fecha_limite, estado, avance)
                VALUES (%s, %s, %s, %s, %s, 'pendiente', 0) RETURNING id
            """, (nombre, descripcion, tipo, prioridad, fecha_limite))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            flash('Proceso creado correctamente', 'success')
        except Exception as e:
            logger.error(f"Error al crear proceso: {e}")
            flash(f'Error al crear el proceso: {str(e)}', 'error')
        
        return redirect(url_for('rh_procesos'))
    
    # GET - Mostrar lista de procesos
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        
        # Obtener filtros
        busqueda = request.args.get('q', '')
        tipo_filtro = request.args.get('tipo', '')
        estado_filtro = request.args.get('estado', '')
        
        # Consulta simplificada - sin JOIN con empleados por ahora
        query = """
            SELECT 
                p.id, 
                p.nombre, 
                p.descripcion,
                p.tipo,
                p.prioridad,
                p.estado,
                p.fecha_limite,
                p.avance,
                p.fecha_inicio,
                p.fecha_completado
            FROM rh_procesos p
            WHERE 1=1
        """
        params = []
        
        if busqueda:
            query += " AND (p.nombre ILIKE %s OR p.descripcion ILIKE %s)"
            params.extend([f'%{busqueda}%', f'%{busqueda}%'])
        
        if tipo_filtro:
            query += " AND p.tipo = %s"
            params.append(tipo_filtro)
        
        if estado_filtro:
            query += " AND p.estado = %s"
            params.append(estado_filtro)
        
        query += " ORDER BY p.id DESC"
        
        cursor.execute(query, params)
        procesos = cursor.fetchall()
        
        # Calcular estadísticas
        try:
            cursor.execute("""
                SELECT 
                    SUM(CASE WHEN estado = 'pendiente' THEN 1 ELSE 0 END) as pendientes,
                    SUM(CASE WHEN estado = 'en_proceso' THEN 1 ELSE 0 END) as en_proceso,
                    SUM(CASE WHEN estado = 'completado' THEN 1 ELSE 0 END) as completados,
                    SUM(CASE WHEN estado = 'cancelado' THEN 1 ELSE 0 END) as cancelados
                FROM rh_procesos
            """)
            stats_row = cursor.fetchone()
            stats = {
                'pendientes': stats_row[0] if stats_row and stats_row[0] else 0,
                'en_proceso': stats_row[1] if stats_row and stats_row[1] else 0,
                'completados': stats_row[2] if stats_row and stats_row[2] else 0,
                'cancelados': stats_row[3] if stats_row and stats_row[3] else 0
            }
        except Exception as e:
            logger.warning(f"Error al calcular estadísticas: {e}")
            stats = {'pendientes': 0, 'en_proceso': 0, 'completados': 0, 'cancelados': 0}
        
        cursor.close()
        conn.close()
        
        return render_template('rh/rh_procesos.html', procesos=procesos, stats=stats)
        
    except Exception as e:
        logger.error(f"Error en rh_procesos: {e}")
        flash(f'Error al cargar los procesos: {str(e)}', 'error')
        return render_template('rh/rh_procesos.html', procesos=[], stats={'pendientes': 0, 'en_proceso': 0, 'completados': 0, 'cancelados': 0})

@app.route('/rh/proceso/<int:id>')
@login_required
def rh_proceso_detalle(id):
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT 
                p.id, 
                p.nombre, 
                p.descripcion,
                p.tipo,
                p.prioridad,
                p.estado,
                p.fecha_limite,
                p.avance,
                p.fecha_inicio,
                p.fecha_completado,
                p.observaciones
            FROM rh_procesos p
            WHERE p.id = %s
        """, (id,))
        
        proceso = cursor.fetchone()
        
        if not proceso:
            flash('Proceso no encontrado', 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('rh_procesos'))
        
        cursor.close()
        conn.close()
        
        return render_template('rh/rh_proceso_detalle.html', proceso=proceso)
        
    except Exception as e:
        logger.error(f"Error en rh_proceso_detalle: {e}")
        flash('Error al cargar el proceso', 'error')
        return redirect(url_for('rh_procesos'))

@app.route('/rh/proceso/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def rh_proceso_editar(id):
    if not current_user.puede('gestionar_rh'):
        flash('No tienes permisos para editar procesos', 'error')
        return redirect(url_for('rh_procesos'))
    
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        
        if request.method == 'POST':
            nombre = request.form.get('nombre')
            descripcion = request.form.get('descripcion')
            estado = request.form.get('estado')
            prioridad = request.form.get('prioridad')
            avance = request.form.get('avance', 0)
            fecha_limite = request.form.get('fecha_limite') or None
            
            # Validar campos obligatorios
            if not nombre:
                flash('El nombre del proceso es obligatorio', 'error')
                cursor.close()
                conn.close()
                return redirect(url_for('rh_proceso_editar', id=id))
            
            # Actualizar proceso
            cursor.execute("""
                UPDATE rh_procesos 
                SET nombre = %s, 
                    descripcion = %s, 
                    estado = %s, 
                    prioridad = %s, 
                    avance = %s, 
                    fecha_limite = %s,
                    fecha_completado = CASE WHEN %s = 'completado' AND estado != 'completado' THEN CURRENT_DATE ELSE fecha_completado END,
                    fecha_inicio = CASE WHEN %s = 'en_proceso' AND estado = 'pendiente' THEN CURRENT_DATE ELSE fecha_inicio END
                WHERE id = %s
            """, (nombre, descripcion, estado, prioridad, avance, fecha_limite, estado, estado, id))
            
            conn.commit()
            flash('Proceso actualizado correctamente', 'success')
            cursor.close()
            conn.close()
            return redirect(url_for('rh_proceso_detalle', id=id))
        
        # GET - Obtener datos del proceso
        cursor.execute("SELECT * FROM rh_procesos WHERE id = %s", (id,))
        proceso = cursor.fetchone()
        
        if not proceso:
            flash('Proceso no encontrado', 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('rh_procesos'))
        
        cursor.close()
        conn.close()
        
        return render_template('rh/rh_proceso_editar.html', proceso=proceso)
        
    except Exception as e:
        logger.error(f"Error en rh_proceso_editar: {e}")
        flash(f'Error al editar el proceso: {str(e)}', 'error')
        return redirect(url_for('rh_procesos'))

@app.route('/rh/capacitaciones')
@login_required
def rh_capacitaciones():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, titulo, tipo, fecha_inicio, fecha_fin, duracion_horas, instructor, estado
            FROM rh_capacitaciones ORDER BY fecha_inicio DESC
        """)
        capacitaciones = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('rh/rh_capacitaciones.html', capacitaciones=capacitaciones)
    except Exception as e:
        logger.error(f"Error en rh_capacitaciones: {e}")
        return render_template('rh/rh_capacitaciones.html', capacitaciones=[])
        
@app.route('/rh/capacitacion/<int:id>')
@login_required
def rh_capacitacion_detalle(id):
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM rh_capacitaciones WHERE id = %s", (id,))
        capacitacion = cursor.fetchone()
        if not capacitacion:
            flash('Capacitación no encontrada', 'error')
            return redirect(url_for('rh_capacitaciones'))
        cursor.execute("""
            SELECT cp.*, e.primer_nombre, e.primer_apellido, e.cargo
            FROM rh_capacitacion_participantes cp
            JOIN rh_empleados e ON cp.empleado_id = e.id
            WHERE cp.capacitacion_id = %s
        """, (id,))
        participantes = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('rh/rh_capacitacion_detalle.html',
                             capacitacion=capacitacion, participantes=participantes)
    except Exception as e:
        logger.error(f"Error en rh_capacitacion_detalle: {e}")
        flash('Error al cargar el detalle de la capacitación', 'error')
        return redirect(url_for('rh_capacitaciones'))

@app.route('/rh/capacitacion/nueva', methods=['GET', 'POST'])
@login_required
def rh_capacitacion_nueva():
    if not current_user.puede('gestionar_rh'):
        flash('No tienes permisos para crear capacitaciones', 'error')
        return redirect(url_for('rh_capacitaciones'))
    try:
        if request.method == 'POST':
            conn = crear_conexion()
            cursor = conn.cursor()
            campos = ['nombre', 'descripcion', 'tipo', 'instructor', 'fecha_inicio',
                     'fecha_fin', 'duracion_horas', 'costo', 'ubicacion', 'observaciones']
            valores = [request.form.get(campo, '') for campo in campos]
            placeholders = ', '.join(['%s'] * len(campos))
            cursor.execute(f"INSERT INTO rh_capacitaciones ({', '.join(campos)}) VALUES ({placeholders}) RETURNING id", valores)
            new_id = cursor.fetchone()[0]
            conn.commit()
            cursor.close()
            conn.close()
            flash('Capacitación creada correctamente', 'success')
            return redirect(url_for('rh_capacitacion_detalle', id=new_id))
        return render_template('rh/rh_capacitacion_nueva.html')
    except Exception as e:
        logger.error(f"Error en rh_capacitacion_nueva: {e}")
        flash('Error al crear la capacitación', 'error')
        return redirect(url_for('rh_capacitaciones'))

@app.route('/rh/evaluaciones')
@login_required
def rh_evaluaciones():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT e.id, emp.primer_nombre as empleado_nombre, emp.primer_apellido as empleado_apellido,
                   eva.primer_nombre as evaluador_nombre, eva.primer_apellido as evaluador_apellido,
                   e.periodo, e.puntaje_total, e.fecha_evaluacion, e.estado
            FROM rh_evaluaciones e
            JOIN rh_empleados emp ON e.empleado_id = emp.id
            JOIN rh_empleados eva ON e.evaluador_id = eva.id
            ORDER BY e.fecha_evaluacion DESC
        """)
        evaluaciones = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('rh/rh_evaluaciones.html', evaluaciones=evaluaciones)
    except Exception as e:
        logger.error(f"Error en rh_evaluaciones: {e}")
        return render_template('rh/rh_evaluaciones.html', evaluaciones=[])

@app.route('/rh/evaluacion/nueva', methods=['GET', 'POST'])
@login_required
def rh_evaluacion_nueva():
    if not current_user.puede('gestionar_rh'):
        flash('No tienes permisos para crear evaluaciones', 'error')
        return redirect(url_for('rh_evaluaciones'))
    try:
        if request.method == 'POST':
            conn = crear_conexion()
            cursor = conn.cursor()
            p1 = int(request.form.get('puntaje_cumplimiento', 0))
            p2 = int(request.form.get('puntaje_trabajo_equipo', 0))
            p3 = int(request.form.get('puntaje_iniciativa', 0))
            p4 = int(request.form.get('puntaje_calidad', 0))
            puntaje_total = round((p1 + p2 + p3 + p4) / 4)
            campos = ['empleado_id', 'evaluador_id', 'periodo', 'fecha_evaluacion',
                     'puntaje_cumplimiento', 'puntaje_trabajo_equipo', 'puntaje_iniciativa',
                     'puntaje_calidad', 'puntaje_total', 'observaciones', 'plan_mejora', 'estado']
            valores = [
                request.form.get('empleado_id'), request.form.get('evaluador_id'),
                request.form.get('periodo'), request.form.get('fecha_evaluacion'),
                p1, p2, p3, p4, puntaje_total,
                request.form.get('observaciones'), request.form.get('plan_mejora'), 'completada'
            ]
            placeholders = ', '.join(['%s'] * len(campos))
            cursor.execute(f"INSERT INTO rh_evaluaciones ({', '.join(campos)}) VALUES ({placeholders}) RETURNING id", valores)
            new_id = cursor.fetchone()[0]
            conn.commit()
            cursor.close()
            conn.close()
            flash('Evaluación creada correctamente', 'success')
            return redirect(url_for('rh_evaluaciones'))
        return render_template('rh/rh_evaluacion_nueva.html')
    except Exception as e:
        logger.error(f"Error en rh_evaluacion_nueva: {e}")
        flash('Error al crear la evaluación', 'error')
        return redirect(url_for('rh_evaluaciones'))

@app.route('/rh/ascensos')
@login_required
def rh_ascensos():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT a.id, e.primer_nombre, e.primer_apellido, c_ant.nombre as cargo_anterior,
                   c_nue.nombre as cargo_nuevo, a.fecha_ascenso, a.salario_anterior,
                   a.salario_nuevo, a.estado
            FROM rh_ascensos a
            JOIN rh_empleados e ON a.empleado_id = e.id
            JOIN rh_cargos c_ant ON a.cargo_anterior_id = c_ant.id
            JOIN rh_cargos c_nue ON a.cargo_nuevo_id = c_nue.id
            ORDER BY a.fecha_ascenso DESC
        """)
        ascensos = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('rh/rh_ascensos.html', ascensos=ascensos)
    except Exception as e:
        logger.error(f"Error en rh_ascensos: {e}")
        return render_template('rh/rh_ascensos.html', ascensos=[])

@app.route('/rh/cargos')
@login_required
def rh_cargos():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT c.*, COUNT(e.id) as num_empleados
            FROM rh_cargos c
            LEFT JOIN rh_empleados e ON c.id = e.cargo_id AND e.estado = 'activo'
            GROUP BY c.id ORDER BY c.nombre
        """)
        cargos = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('rh/rh_cargos.html', cargos=cargos)
    except Exception as e:
        logger.error(f"Error en rh_cargos: {e}")
        return render_template('rh/rh_cargos.html', cargos=[])

@app.route('/rh/cargo/nuevo', methods=['GET', 'POST'])
@login_required
def rh_cargo_nuevo():
    if not current_user.puede('gestionar_rh'):
        flash('No tienes permisos para crear cargos', 'error')
        return redirect(url_for('rh_cargos'))
    try:
        if request.method == 'POST':
            conn = crear_conexion()
            cursor = conn.cursor()
            campos = ['nombre', 'descripcion', 'departamento', 'nivel', 'salario_base',
                     'rango_min', 'rango_max', 'requisitos', 'competencias']
            valores = [request.form.get(campo, '') for campo in campos]
            placeholders = ', '.join(['%s'] * len(campos))
            cursor.execute(f"INSERT INTO rh_cargos ({', '.join(campos)}) VALUES ({placeholders}) RETURNING id", valores)
            new_id = cursor.fetchone()[0]
            conn.commit()
            cursor.close()
            conn.close()
            flash('Cargo creado correctamente', 'success')
            return redirect(url_for('rh_cargos'))
        return render_template('rh/rh_cargo_nuevo.html')
    except Exception as e:
        logger.error(f"Error en rh_cargo_nuevo: {e}")
        flash('Error al crear el cargo', 'error')
        return redirect(url_for('rh_cargos'))

@app.route('/rh/sanciones')
@login_required
def rh_sanciones():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        
        # Consulta CORREGIDA - usando las columnas correctas de tu BD
        cursor.execute("""
            SELECT 
                s.id,
                s.empleado_id,
                s.tipo,
                s.fecha,
                s.motivo,
                s.descripcion,
                s.duracion_dias,
                s.documento,
                s.estado,
                s.created_at,
                e.primer_nombre,
                e.primer_apellido,
                e.documento as documento_empleado,
                COALESCE(c.nombre, 'Sin asignar') as cargo
            FROM rh_sanciones s 
            JOIN rh_empleados e ON s.empleado_id = e.id
            LEFT JOIN rh_cargos c ON e.cargo_id = c.id
            ORDER BY s.fecha DESC
        """)
        
        sanciones = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return render_template('rh/rh_sanciones.html', sanciones=sanciones)
        
    except Exception as e:
        logger.error(f"Error en rh_sanciones: {e}")
        flash('Error al cargar las sanciones', 'error')
        return render_template('rh/rh_sanciones.html', sanciones=[])

@app.route('/rh/sancion/nueva', methods=['GET', 'POST'])
@login_required
def rh_sancion_nueva():
    if not current_user.puede('gestionar_rh'):
        flash('No tienes permisos para registrar sanciones', 'error')
        return redirect(url_for('rh_sanciones'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        
        # Obtener empleados para el select
        cursor.execute("""
            SELECT e.id, e.primer_nombre, e.primer_apellido, e.documento
            FROM rh_empleados e
            WHERE e.estado = 'activo'
            ORDER BY e.primer_apellido
        """)
        empleados = cursor.fetchall()
        
        if request.method == 'POST':
            empleado_id = request.form.get('empleado_id')
            tipo = request.form.get('tipo')
            fecha = request.form.get('fecha')
            motivo = request.form.get('motivo')
            duracion_dias = request.form.get('duracion_dias') or None
            descripcion = request.form.get('descripcion')
            
            if not empleado_id or not tipo or not fecha or not motivo:
                flash('❌ Todos los campos obligatorios deben ser diligenciados', 'error')
                return render_template('rh/rh_sancion_nueva.html', empleados=empleados)
            
            cursor.execute("""
                INSERT INTO rh_sanciones (empleado_id, tipo, fecha, motivo, descripcion, duracion_dias, estado)
                VALUES (%s, %s, %s, %s, %s, %s, 'vigente') RETURNING id
            """, (empleado_id, tipo, fecha, motivo, descripcion, duracion_dias))
            
            new_id = cursor.fetchone()[0]
            conn.commit()
            cursor.close()
            conn.close()
            
            flash('✅ Sanción registrada correctamente', 'success')
            return redirect(url_for('rh_sanciones'))
        
        cursor.close()
        conn.close()
        return render_template('rh/rh_sancion_nueva.html', empleados=empleados)
        
    except Exception as e:
        logger.error(f"Error en rh_sancion_nueva: {e}")
        flash('Error al registrar la sanción', 'error')
        return redirect(url_for('rh_sanciones'))

@app.route('/rh/departamentos')
@login_required
def rh_departamentos():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT d.*, COUNT(e.id) as num_empleados
            FROM rh_departamentos d
            LEFT JOIN rh_empleados e ON d.id = e.departamento_id AND e.estado = 'activo'
            GROUP BY d.id ORDER BY d.nombre
        """)
        departamentos = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('rh/rh_departamentos.html', departamentos=departamentos)
    except Exception as e:
        logger.error(f"Error en rh_departamentos: {e}")
        return render_template('rh/rh_departamentos.html', departamentos=[])

@app.route('/rh/organigrama')
@login_required
def rh_organigrama():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    return render_template('rh/rh_organigrama.html')

@app.route('/rh/reportes')
@login_required
def rh_reportes():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    return render_template('rh/rh_reportes.html')

@app.route('/rh/reporte/asistencia')
@login_required
def rh_reporte_asistencia():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    return render_template('rh/rh_reporte_asistencia.html')

@app.route('/rh/reporte/nomina')
@login_required
def rh_reporte_nomina():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    return render_template('rh/rh_reporte_nomina.html')

@app.route('/rh/reporte/rotacion')
@login_required
def rh_reporte_rotacion():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    return render_template('rh/rh_reporte_rotacion.html')

@app.route('/rh/normativas')
@login_required
def rh_normativas():
    if not current_user.puede('ver_rh'):
        flash('No tienes permisos para acceder a Recursos Humanos', 'error')
        return redirect(url_for('index'))
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT id, titulo, descripcion, tipo, fecha_publicacion, url
                FROM rh_normativas ORDER BY fecha_publicacion DESC
            """)
            normativas = cursor.fetchall()
        except Exception as e:
            logger.warning(f"Tabla rh_normativas no existe: {e}")
            normativas = []
        cursor.close()
        conn.close()
        return render_template('rh/rh_normativas.html', normativas=normativas)
    except Exception as e:
        logger.error(f"Error en rh_normativas: {e}")
        flash('Error al cargar las normativas', 'error')
        return redirect(url_for('rh_dashboard'))

# ==========================================
# FIN MÓDULO DE RECURSOS HUMANOS
# ==========================================

# ===== DASHBOARD PRINCIPAL DEL ADMINISTRADOR =====
@app.route('/dashboard_admin')
@login_required
@admin_required
def dashboard_admin():
    """
    Panel HUB del administrador: muestra tarjetas para ingresar
    a Soporte, SST y RH. También muestra estadísticas generales.
    """
    try:
        conn = crear_conexion()
        cursor = conn.cursor()

        def safe_count(query):
            try:
                cursor.execute(query)
                return cursor.fetchone()[0]
            except Exception:
                return 0

        stats = {
            'total_usuarios':          safe_count("SELECT COUNT(*) FROM usuarios"),
            'usuarios_soporte':        safe_count("SELECT COUNT(*) FROM usuarios WHERE rol IN ('soporte', 'usuario')"),
            'usuarios_sst':            safe_count("SELECT COUNT(*) FROM usuarios WHERE rol = 'sst'"),
            'usuarios_rh':             safe_count("SELECT COUNT(*) FROM usuarios WHERE rol = 'rh'"),
            'total_fichas':            safe_count("SELECT COUNT(*) FROM fichas"),
            'total_contenido_sst':     safe_count("SELECT COUNT(*) FROM sst_contenido"),
            'total_actividades_plan':  safe_count("SELECT COUNT(*) FROM plan_anual_trabajo"),
            'total_empleados_activos': safe_count("SELECT COUNT(*) FROM rh_empleados WHERE estado = 'activo'"),
            'procesos_pendientes': 0,
        }

        cursor.close()
        conn.close()

        urls = {
            'gestion_usuarios':   url_for('gestion_usuarios'),
            'agregar_usuario':    url_for('agregar_usuario'),
            'cambiar_password':   url_for('cambiar_password'),
            'soluciones_visuales': url_for('soluciones_visuales'),
            'atencion_telefonica': url_for('atencion_telefonica'),
            'informacion_general': url_for('informacion_general'),
            'logout':             url_for('logout'),
            'index':              url_for('index'),
            'sst_dashboard':      url_for('sst_dashboard'),
            'rh_dashboard':       url_for('rh_dashboard'),
        }

        return render_template('dashboard_admin.html', stats=stats, urls=urls)

    except Exception as e:
        logger.error(f"Error en dashboard_admin: {e}")
        return render_template('dashboard_admin.html', stats={}, urls={})


# ===== FUNCIONES AUXILIARES =====
def actualizar_porcentaje_avance(id):
    """Actualizar automáticamente el porcentaje de avance de una actividad"""
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
        columnas_p = [f'{mes}_semana{s}_p' for mes in meses for s in range(1, 5)]
        columnas_e = [col.replace('_p', '_e') for col in columnas_p]
        query_p = f"SELECT {' + '.join([f'CASE WHEN {col} = TRUE THEN 1 ELSE 0 END' for col in columnas_p])} FROM plan_anual_trabajo WHERE id = %s"
        cursor.execute(query_p, (id,))
        total_planificadas = cursor.fetchone()[0]
        query_e = f"SELECT {' + '.join([f'CASE WHEN {col} = TRUE THEN 1 ELSE 0 END' for col in columnas_e])} FROM plan_anual_trabajo WHERE id = %s"
        cursor.execute(query_e, (id,))
        total_ejecutadas = cursor.fetchone()[0]
        porcentaje = 0
        if total_planificadas > 0:
            porcentaje = round((total_ejecutadas / total_planificadas) * 100, 2)
        if porcentaje == 100:
            estado = 'completado'
        elif porcentaje > 0:
            estado = 'en_proceso'
        else:
            estado = 'pendiente'
        cursor.execute("UPDATE plan_anual_trabajo SET porcentaje_avance = %s, estado = %s WHERE id = %s", (porcentaje, estado, id))
        conn.commit()
        cursor.close()
        conn.close()
        logger.info(f"✅ Porcentaje actualizado para actividad {id}: {porcentaje}% ({estado})")
        return porcentaje, estado
    except Exception as e:
        logger.error(f"❌ Error al actualizar porcentaje: {e}")
        return 0, 'pendiente'


def inicializar_plan_anual():
    """Crear tablas del plan anual si no existen"""
    try:
        conn = crear_conexion()
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS plan_anual_trabajo (
                id SERIAL PRIMARY KEY,
                actividad TEXT NOT NULL,
                evidencia TEXT,
                ciclo_phva VARCHAR(50),
                articulos_decreto VARCHAR(200),
                nivel_pesv VARCHAR(100),
                responsables VARCHAR(200),
                recursos TEXT,
                enero_semana1_p BOOLEAN DEFAULT FALSE, enero_semana1_e BOOLEAN DEFAULT FALSE,
                enero_semana2_p BOOLEAN DEFAULT FALSE, enero_semana2_e BOOLEAN DEFAULT FALSE,
                enero_semana3_p BOOLEAN DEFAULT FALSE, enero_semana3_e BOOLEAN DEFAULT FALSE,
                enero_semana4_p BOOLEAN DEFAULT FALSE, enero_semana4_e BOOLEAN DEFAULT FALSE,
                febrero_semana1_p BOOLEAN DEFAULT FALSE, febrero_semana1_e BOOLEAN DEFAULT FALSE,
                febrero_semana2_p BOOLEAN DEFAULT FALSE, febrero_semana2_e BOOLEAN DEFAULT FALSE,
                febrero_semana3_p BOOLEAN DEFAULT FALSE, febrero_semana3_e BOOLEAN DEFAULT FALSE,
                febrero_semana4_p BOOLEAN DEFAULT FALSE, febrero_semana4_e BOOLEAN DEFAULT FALSE,
                marzo_semana1_p BOOLEAN DEFAULT FALSE, marzo_semana1_e BOOLEAN DEFAULT FALSE,
                marzo_semana2_p BOOLEAN DEFAULT FALSE, marzo_semana2_e BOOLEAN DEFAULT FALSE,
                marzo_semana3_p BOOLEAN DEFAULT FALSE, marzo_semana3_e BOOLEAN DEFAULT FALSE,
                marzo_semana4_p BOOLEAN DEFAULT FALSE, marzo_semana4_e BOOLEAN DEFAULT FALSE,
                abril_semana1_p BOOLEAN DEFAULT FALSE, abril_semana1_e BOOLEAN DEFAULT FALSE,
                abril_semana2_p BOOLEAN DEFAULT FALSE, abril_semana2_e BOOLEAN DEFAULT FALSE,
                abril_semana3_p BOOLEAN DEFAULT FALSE, abril_semana3_e BOOLEAN DEFAULT FALSE,
                abril_semana4_p BOOLEAN DEFAULT FALSE, abril_semana4_e BOOLEAN DEFAULT FALSE,
                mayo_semana1_p BOOLEAN DEFAULT FALSE, mayo_semana1_e BOOLEAN DEFAULT FALSE,
                mayo_semana2_p BOOLEAN DEFAULT FALSE, mayo_semana2_e BOOLEAN DEFAULT FALSE,
                mayo_semana3_p BOOLEAN DEFAULT FALSE, mayo_semana3_e BOOLEAN DEFAULT FALSE,
                mayo_semana4_p BOOLEAN DEFAULT FALSE, mayo_semana4_e BOOLEAN DEFAULT FALSE,
                junio_semana1_p BOOLEAN DEFAULT FALSE, junio_semana1_e BOOLEAN DEFAULT FALSE,
                junio_semana2_p BOOLEAN DEFAULT FALSE, junio_semana2_e BOOLEAN DEFAULT FALSE,
                junio_semana3_p BOOLEAN DEFAULT FALSE, junio_semana3_e BOOLEAN DEFAULT FALSE,
                junio_semana4_p BOOLEAN DEFAULT FALSE, junio_semana4_e BOOLEAN DEFAULT FALSE,
                julio_semana1_p BOOLEAN DEFAULT FALSE, julio_semana1_e BOOLEAN DEFAULT FALSE,
                julio_semana2_p BOOLEAN DEFAULT FALSE, julio_semana2_e BOOLEAN DEFAULT FALSE,
                julio_semana3_p BOOLEAN DEFAULT FALSE, julio_semana3_e BOOLEAN DEFAULT FALSE,
                julio_semana4_p BOOLEAN DEFAULT FALSE, julio_semana4_e BOOLEAN DEFAULT FALSE,
                agosto_semana1_p BOOLEAN DEFAULT FALSE, agosto_semana1_e BOOLEAN DEFAULT FALSE,
                agosto_semana2_p BOOLEAN DEFAULT FALSE, agosto_semana2_e BOOLEAN DEFAULT FALSE,
                agosto_semana3_p BOOLEAN DEFAULT FALSE, agosto_semana3_e BOOLEAN DEFAULT FALSE,
                agosto_semana4_p BOOLEAN DEFAULT FALSE, agosto_semana4_e BOOLEAN DEFAULT FALSE,
                septiembre_semana1_p BOOLEAN DEFAULT FALSE, septiembre_semana1_e BOOLEAN DEFAULT FALSE,
                septiembre_semana2_p BOOLEAN DEFAULT FALSE, septiembre_semana2_e BOOLEAN DEFAULT FALSE,
                septiembre_semana3_p BOOLEAN DEFAULT FALSE, septiembre_semana3_e BOOLEAN DEFAULT FALSE,
                septiembre_semana4_p BOOLEAN DEFAULT FALSE, septiembre_semana4_e BOOLEAN DEFAULT FALSE,
                octubre_semana1_p BOOLEAN DEFAULT FALSE, octubre_semana1_e BOOLEAN DEFAULT FALSE,
                octubre_semana2_p BOOLEAN DEFAULT FALSE, octubre_semana2_e BOOLEAN DEFAULT FALSE,
                octubre_semana3_p BOOLEAN DEFAULT FALSE, octubre_semana3_e BOOLEAN DEFAULT FALSE,
                octubre_semana4_p BOOLEAN DEFAULT FALSE, octubre_semana4_e BOOLEAN DEFAULT FALSE,
                noviembre_semana1_p BOOLEAN DEFAULT FALSE, noviembre_semana1_e BOOLEAN DEFAULT FALSE,
                noviembre_semana2_p BOOLEAN DEFAULT FALSE, noviembre_semana2_e BOOLEAN DEFAULT FALSE,
                noviembre_semana3_p BOOLEAN DEFAULT FALSE, noviembre_semana3_e BOOLEAN DEFAULT FALSE,
                noviembre_semana4_p BOOLEAN DEFAULT FALSE, noviembre_semana4_e BOOLEAN DEFAULT FALSE,
                diciembre_semana1_p BOOLEAN DEFAULT FALSE, diciembre_semana1_e BOOLEAN DEFAULT FALSE,
                diciembre_semana2_p BOOLEAN DEFAULT FALSE, diciembre_semana2_e BOOLEAN DEFAULT FALSE,
                diciembre_semana3_p BOOLEAN DEFAULT FALSE, diciembre_semana3_e BOOLEAN DEFAULT FALSE,
                diciembre_semana4_p BOOLEAN DEFAULT FALSE, diciembre_semana4_e BOOLEAN DEFAULT FALSE,
                observaciones TEXT,
                estado VARCHAR(20) DEFAULT 'pendiente',
                porcentaje_avance DECIMAL(5,2) DEFAULT 0.00,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                fecha_actualizacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                usuario_actualizacion INTEGER
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS plan_evidencias (
                id SERIAL PRIMARY KEY,
                plan_id INTEGER REFERENCES plan_anual_trabajo(id) ON DELETE CASCADE,
                titulo VARCHAR(200) NOT NULL,
                descripcion TEXT,
                archivo_nombre VARCHAR(300),
                archivo_tipo VARCHAR(100),
                archivo_tamano INTEGER,
                archivo_data BYTEA,
                usuario_id INTEGER,
                fecha_carga TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS plan_seguimiento (
                id SERIAL PRIMARY KEY,
                actividad_id INTEGER REFERENCES plan_anual_trabajo(id) ON DELETE CASCADE,
                comentario TEXT NOT NULL,
                tipo VARCHAR(50),
                usuario_id INTEGER,
                fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()
        cursor.close()
        conn.close()
        print("✅ Tablas del plan anual verificadas/creadas")
    except Exception as e:
        print(f"❌ Error al inicializar plan anual: {e}")
        import traceback
        traceback.print_exc()

# ===== CREAR USUARIO RH AUTOMÁTICAMENTE AL INICIAR =====
def crear_usuario_rh_si_no_existe():
    """Crea el usuario RH automáticamente si no existe"""
    try:
        from werkzeug.security import generate_password_hash
        
        # Verificar si existe
        resultado = ejecutar_consulta(
            "SELECT id FROM usuarios WHERE usuario = %s",
            ('rh',),
            fetch=True
        )
        
        if not resultado:
            # Crear usuario RH
            password_hash = generate_password_hash('rh123')
            ejecutar_consulta("""
                INSERT INTO usuarios (usuario, password, rol, modulo_principal, permisos) 
                VALUES (%s, %s, %s, %s, %s)
            """, (
                'rh', 
                password_hash, 
                'rh', 
                'rh',
                '{"ver_rh": true, "gestionar_rh": true, "cambiar_password": true}'
            ), commit=True)
            print("=" * 50)
            print("✅ Usuario RH creado automáticamente")
            print("👤 Usuario: rh")
            print("🔑 Contraseña: rh123")
            print("=" * 50)
        else:
            print("✅ Usuario RH ya existe")
    except Exception as e:
        print(f"⚠️ No se pudo crear usuario RH: {e}")

# Llamar a la función
crear_usuario_rh_si_no_existe()


# ===== INICIALIZACIÓN =====
if __name__ == '__main__':
    with app.app_context():
        print("🚀 Iniciando la aplicación Flask...")
        print("📊 Creando tablas en la base de datos...")
        crear_tablas()
        print("✅ Tablas creadas/verificadas correctamente")
        print("📋 Verificando categorías SST...")
        try:
            verificar_y_crear_categorias_sst()
            print("✅ Categorías SST verificadas correctamente")
        except Exception as e:
            print(f"⚠️  Advertencia al crear categorías SST: {e}")
        print("📥 Inicializando datos del plan anual...")
        inicializar_plan_anual()
    print("🌐 Aplicación lista en http://0.0.0.0:5000")
    app.run(host='0.0.0.0', port=5000, debug=True)
