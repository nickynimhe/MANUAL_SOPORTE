#!/usr/bin/env python3
"""
Script para importar el Plan Anual de Trabajo 2026 desde Excel a PostgreSQL
Ejecutar este script UNA VEZ después de crear las tablas
"""

import openpyxl
import psycopg2
from config import Config

def importar_plan_anual_desde_excel():
    """Importar todas las actividades del Excel a la base de datos"""
    
    print("📊 INICIANDO IMPORTACIÓN DEL PLAN ANUAL 2026...")
    
    # 1. Conectar a la base de datos
    try:
        conn = psycopg2.connect(
            host=Config.DB_HOST,
            database=Config.DB_NAME,
            user=Config.DB_USER,
            password=Config.DB_PASSWORD,
            port=Config.DB_PORT,
            sslmode='require'
        )
        cursor = conn.cursor()
        print("✅ Conectado a la base de datos")
    except Exception as e:
        print(f"❌ Error al conectar a la BD: {e}")
        return False
    
    # 2. Verificar si ya hay datos
    cursor.execute("SELECT COUNT(*) FROM plan_anual_trabajo")
    count = cursor.fetchone()[0]
    
    if count > 7:  # Si hay más de las 7 actividades de prueba
        respuesta = input(f"\n⚠️  Ya existen {count} actividades en la BD. ¿Deseas eliminarlas y volver a importar? (s/n): ")
        if respuesta.lower() != 's':
            print("❌ Importación cancelada")
            cursor.close()
            conn.close()
            return False
        
        # Eliminar actividades existentes
        cursor.execute("DELETE FROM plan_anual_trabajo")
        conn.commit()
        print("✅ Actividades existentes eliminadas")
    
    # 3. Cargar el Excel
    try:
        wb = openpyxl.load_workbook('/mnt/user-data/uploads/Plan_Anual_de_Trabajo_2026.xlsx')
        ws = wb.active
        print("✅ Excel cargado correctamente")
    except Exception as e:
        print(f"❌ Error al cargar Excel: {e}")
        cursor.close()
        conn.close()
        return False
    
    # 4. Extraer actividades
    actividades = []
    fila_actual = 13  # Los datos empiezan en fila 13
    
    print("📥 Extrayendo actividades del Excel...")
    
    while fila_actual <= ws.max_row:
        # Leer columnas básicas
        actividad = ws.cell(fila_actual, 2).value  # Columna B
        evidencia = ws.cell(fila_actual, 3).value  # Columna C
        ciclo_phva = ws.cell(fila_actual, 4).value  # Columna D
        articulos = ws.cell(fila_actual, 5).value  # Columna E
        nivel_pesv = ws.cell(fila_actual, 6).value  # Columna F
        responsables = ws.cell(fila_actual, 7).value  # Columna G
        recursos = ws.cell(fila_actual, 8).value  # Columna H
        
        # Si no hay actividad, es fila vacía o separador
        if not actividad or isinstance(actividad, str) and (
            'PLANEAR' in actividad.upper() or 
            'HACER' in actividad.upper() or 
            'VERIFICAR' in actividad.upper() or 
            'ACTUAR' in actividad.upper() or
            'DISEÑO' in actividad.upper()
        ):
            fila_actual += 1
            continue
        
        # Leer programación mensual
        programacion = {}
        meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
        
        col_inicio = 9  # Columna I (enero semana 1 P)
        
        for idx, mes in enumerate(meses):
            programacion[mes] = []
            
            # Cada mes tiene 8 columnas (4 semanas x 2 (P/E))
            mes_col_inicio = col_inicio + (idx * 8)
            
            for semana in range(4):
                col_p = mes_col_inicio + (semana * 2)
                col_e = col_p + 1
                
                val_p = ws.cell(fila_actual, col_p).value
                val_e = ws.cell(fila_actual, col_e).value
                
                # Convertir a booleano
                planificado = val_p in ['x', 'X', True, 1, '1'] if val_p else False
                ejecutado = val_e in ['x', 'X', True, 1, '1'] if val_e else False
                
                programacion[mes].append({
                    'planificado': planificado,
                    'ejecutado': ejecutado
                })
        
        actividades.append({
            'actividad': str(actividad).strip() if actividad else '',
            'evidencia': str(evidencia).strip() if evidencia else '',
            'ciclo_phva': str(ciclo_phva).strip() if ciclo_phva else '',
            'articulos': str(articulos).strip() if articulos else '',
            'nivel_pesv': str(nivel_pesv).strip() if nivel_pesv else '',
            'responsables': str(responsables).strip() if responsables else '',
            'recursos': str(recursos).strip() if recursos else '',
            'programacion': programacion
        })
        
        fila_actual += 1
    
    print(f"✅ {len(actividades)} actividades extraídas del Excel")
    
    # 5. Insertar en la base de datos
    print("💾 Insertando actividades en la base de datos...")
    
    meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
            'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
    
    insertadas = 0
    errores = 0
    
    for act in actividades:
        try:
            # Construir columnas y valores
            columnas = ['actividad', 'evidencia', 'ciclo_phva', 'articulos_decreto',
                       'nivel_pesv', 'responsables', 'recursos', 'estado']
            valores = [
                act['actividad'][:500] if act['actividad'] else None,  # Limitar a 500 chars
                act['evidencia'][:500] if act['evidencia'] else None,
                act['ciclo_phva'][:50] if act['ciclo_phva'] else None,
                act['articulos'][:200] if act['articulos'] else None,
                act['nivel_pesv'][:100] if act['nivel_pesv'] else None,
                act['responsables'][:200] if act['responsables'] else None,
                act['recursos'][:200] if act['recursos'] else None,
                'pendiente'  # Estado inicial
            ]
            
            # Agregar programación mensual
            for mes in meses:
                if mes in act['programacion']:
                    semanas = act['programacion'][mes]
                    for semana_idx, semana in enumerate(semanas, 1):
                        columnas.append(f'{mes}_semana{semana_idx}_p')
                        valores.append(semana['planificado'])
                        
                        columnas.append(f'{mes}_semana{semana_idx}_e')
                        valores.append(semana['ejecutado'])
                else:
                    # Si no hay programación para este mes, llenar con False
                    for semana in range(1, 5):
                        columnas.append(f'{mes}_semana{semana}_p')
                        valores.append(False)
                        
                        columnas.append(f'{mes}_semana{semana}_e')
                        valores.append(False)
            
            # Crear query
            placeholders = ', '.join(['%s'] * len(valores))
            query = f"""
                INSERT INTO plan_anual_trabajo ({', '.join(columnas)})
                VALUES ({placeholders})
            """
            
            cursor.execute(query, valores)
            insertadas += 1
            
            # Mostrar progreso cada 10 actividades
            if insertadas % 10 == 0:
                print(f"  → {insertadas} actividades insertadas...")
            
        except Exception as e:
            errores += 1
            print(f"  ⚠️  Error insertando actividad '{act['actividad'][:50]}...': {e}")
    
    # 6. Commit y cerrar
    conn.commit()
    cursor.close()
    conn.close()
    
    print(f"\n✅ IMPORTACIÓN COMPLETADA:")
    print(f"   • {insertadas} actividades insertadas correctamente")
    print(f"   • {errores} errores")
    
    return True


if __name__ == '__main__':
    print("""
╔══════════════════════════════════════════════════════════╗
║  IMPORTADOR DE PLAN ANUAL DE TRABAJO 2026               ║
║  M@STV Producciones - Sistema de Gestión SST            ║
╚══════════════════════════════════════════════════════════╝
    """)
    
    importar_plan_anual_desde_excel()
    
    print("\n✅ Proceso finalizado. Puedes cerrar esta ventana.")
