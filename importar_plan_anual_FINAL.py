#!/usr/bin/env python3
"""
Script de importación CORREGIDO del Plan Anual de Trabajo PESV 2026
Ejecutar en el servidor con: python3 importar_plan_anual_FINAL.py
"""

import openpyxl
import psycopg2
import sys
import os

# Configuración de conexión (usar las mismas credenciales de config.py)
DB_CONFIG = {
    'host': 'dpg-d4g6i23e5dus739l1c80-a.oregon-postgres.render.com',
    'database': 'soporte_tecnico_bujd',
    'user': 'soporte_tecnico_bujd_user',
    'password': '4O43zJ3NiE5NrvdeMYD3hxsXgIOWVonw',
    'port': '5432',
    'sslmode': 'require'
}

# Ruta del archivo Excel (ajustar si es necesario)
EXCEL_PATH = 'Plan_Anual_de_Trabajo_2026.xlsx'

def crear_tabla_si_no_existe(cursor):
    """Crear la tabla si no existe - CON TODAS LAS COLUMNAS NECESARIAS"""
    print("🔧 Verificando/creando tabla plan_anual_trabajo...")
    
    create_table_query = """
        CREATE TABLE IF NOT EXISTS plan_anual_trabajo (
            id SERIAL PRIMARY KEY,
            actividad VARCHAR(500) NOT NULL,
            evidencia VARCHAR(500),
            ciclo_phva VARCHAR(50),
            articulos_decreto VARCHAR(200),
            nivel_pesv VARCHAR(100),
            responsables VARCHAR(200),
            recursos VARCHAR(200),
            -- Enero
            enero_semana1_p BOOLEAN DEFAULT FALSE,
            enero_semana1_e BOOLEAN DEFAULT FALSE,
            enero_semana2_p BOOLEAN DEFAULT FALSE,
            enero_semana2_e BOOLEAN DEFAULT FALSE,
            enero_semana3_p BOOLEAN DEFAULT FALSE,
            enero_semana3_e BOOLEAN DEFAULT FALSE,
            enero_semana4_p BOOLEAN DEFAULT FALSE,
            enero_semana4_e BOOLEAN DEFAULT FALSE,
            -- Febrero
            febrero_semana1_p BOOLEAN DEFAULT FALSE,
            febrero_semana1_e BOOLEAN DEFAULT FALSE,
            febrero_semana2_p BOOLEAN DEFAULT FALSE,
            febrero_semana2_e BOOLEAN DEFAULT FALSE,
            febrero_semana3_p BOOLEAN DEFAULT FALSE,
            febrero_semana3_e BOOLEAN DEFAULT FALSE,
            febrero_semana4_p BOOLEAN DEFAULT FALSE,
            febrero_semana4_e BOOLEAN DEFAULT FALSE,
            -- Marzo
            marzo_semana1_p BOOLEAN DEFAULT FALSE,
            marzo_semana1_e BOOLEAN DEFAULT FALSE,
            marzo_semana2_p BOOLEAN DEFAULT FALSE,
            marzo_semana2_e BOOLEAN DEFAULT FALSE,
            marzo_semana3_p BOOLEAN DEFAULT FALSE,
            marzo_semana3_e BOOLEAN DEFAULT FALSE,
            marzo_semana4_p BOOLEAN DEFAULT FALSE,
            marzo_semana4_e BOOLEAN DEFAULT FALSE,
            -- Abril
            abril_semana1_p BOOLEAN DEFAULT FALSE,
            abril_semana1_e BOOLEAN DEFAULT FALSE,
            abril_semana2_p BOOLEAN DEFAULT FALSE,
            abril_semana2_e BOOLEAN DEFAULT FALSE,
            abril_semana3_p BOOLEAN DEFAULT FALSE,
            abril_semana3_e BOOLEAN DEFAULT FALSE,
            abril_semana4_p BOOLEAN DEFAULT FALSE,
            abril_semana4_e BOOLEAN DEFAULT FALSE,
            -- Mayo
            mayo_semana1_p BOOLEAN DEFAULT FALSE,
            mayo_semana1_e BOOLEAN DEFAULT FALSE,
            mayo_semana2_p BOOLEAN DEFAULT FALSE,
            mayo_semana2_e BOOLEAN DEFAULT FALSE,
            mayo_semana3_p BOOLEAN DEFAULT FALSE,
            mayo_semana3_e BOOLEAN DEFAULT FALSE,
            mayo_semana4_p BOOLEAN DEFAULT FALSE,
            mayo_semana4_e BOOLEAN DEFAULT FALSE,
            -- Junio
            junio_semana1_p BOOLEAN DEFAULT FALSE,
            junio_semana1_e BOOLEAN DEFAULT FALSE,
            junio_semana2_p BOOLEAN DEFAULT FALSE,
            junio_semana2_e BOOLEAN DEFAULT FALSE,
            junio_semana3_p BOOLEAN DEFAULT FALSE,
            junio_semana3_e BOOLEAN DEFAULT FALSE,
            junio_semana4_p BOOLEAN DEFAULT FALSE,
            junio_semana4_e BOOLEAN DEFAULT FALSE,
            -- Julio
            julio_semana1_p BOOLEAN DEFAULT FALSE,
            julio_semana1_e BOOLEAN DEFAULT FALSE,
            julio_semana2_p BOOLEAN DEFAULT FALSE,
            julio_semana2_e BOOLEAN DEFAULT FALSE,
            julio_semana3_p BOOLEAN DEFAULT FALSE,
            julio_semana3_e BOOLEAN DEFAULT FALSE,
            julio_semana4_p BOOLEAN DEFAULT FALSE,
            julio_semana4_e BOOLEAN DEFAULT FALSE,
            -- Agosto
            agosto_semana1_p BOOLEAN DEFAULT FALSE,
            agosto_semana1_e BOOLEAN DEFAULT FALSE,
            agosto_semana2_p BOOLEAN DEFAULT FALSE,
            agosto_semana2_e BOOLEAN DEFAULT FALSE,
            agosto_semana3_p BOOLEAN DEFAULT FALSE,
            agosto_semana3_e BOOLEAN DEFAULT FALSE,
            agosto_semana4_p BOOLEAN DEFAULT FALSE,
            agosto_semana4_e BOOLEAN DEFAULT FALSE,
            -- Septiembre
            septiembre_semana1_p BOOLEAN DEFAULT FALSE,
            septiembre_semana1_e BOOLEAN DEFAULT FALSE,
            septiembre_semana2_p BOOLEAN DEFAULT FALSE,
            septiembre_semana2_e BOOLEAN DEFAULT FALSE,
            septiembre_semana3_p BOOLEAN DEFAULT FALSE,
            septiembre_semana3_e BOOLEAN DEFAULT FALSE,
            septiembre_semana4_p BOOLEAN DEFAULT FALSE,
            septiembre_semana4_e BOOLEAN DEFAULT FALSE,
            -- Octubre
            octubre_semana1_p BOOLEAN DEFAULT FALSE,
            octubre_semana1_e BOOLEAN DEFAULT FALSE,
            octubre_semana2_p BOOLEAN DEFAULT FALSE,
            octubre_semana2_e BOOLEAN DEFAULT FALSE,
            octubre_semana3_p BOOLEAN DEFAULT FALSE,
            octubre_semana3_e BOOLEAN DEFAULT FALSE,
            octubre_semana4_p BOOLEAN DEFAULT FALSE,
            octubre_semana4_e BOOLEAN DEFAULT FALSE,
            -- Noviembre
            noviembre_semana1_p BOOLEAN DEFAULT FALSE,
            noviembre_semana1_e BOOLEAN DEFAULT FALSE,
            noviembre_semana2_p BOOLEAN DEFAULT FALSE,
            noviembre_semana2_e BOOLEAN DEFAULT FALSE,
            noviembre_semana3_p BOOLEAN DEFAULT FALSE,
            noviembre_semana3_e BOOLEAN DEFAULT FALSE,
            noviembre_semana4_p BOOLEAN DEFAULT FALSE,
            noviembre_semana4_e BOOLEAN DEFAULT FALSE,
            -- Diciembre
            diciembre_semana1_p BOOLEAN DEFAULT FALSE,
            diciembre_semana1_e BOOLEAN DEFAULT FALSE,
            diciembre_semana2_p BOOLEAN DEFAULT FALSE,
            diciembre_semana2_e BOOLEAN DEFAULT FALSE,
            diciembre_semana3_p BOOLEAN DEFAULT FALSE,
            diciembre_semana3_e BOOLEAN DEFAULT FALSE,
            diciembre_semana4_p BOOLEAN DEFAULT FALSE,
            diciembre_semana4_e BOOLEAN DEFAULT FALSE,
            -- Campos de control
            estado VARCHAR(20) DEFAULT 'pendiente',
            porcentaje_avance DECIMAL(5,2) DEFAULT 0.00,
            fecha_actualizacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            usuario_actualizacion INTEGER,
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """
    
    try:
        cursor.execute(create_table_query)
        print("✅ Tabla 'plan_anual_trabajo' creada/verificada")
        return True
    except Exception as e:
        print(f"❌ Error al crear tabla: {e}")
        return False

def main():
    print("=" * 60)
    print("IMPORTACIÓN CORREGIDA DEL PLAN ANUAL DE TRABAJO PESV 2026")
    print("=" * 60)
    
    # Verificar que el archivo existe
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Error: No se encuentra el archivo {EXCEL_PATH}")
        print("   Asegúrate de ejecutar este script en la misma carpeta que el Excel")
        print("   Ruta actual:", os.getcwd())
        sys.exit(1)
    else:
        print(f"✅ Archivo encontrado: {EXCEL_PATH}")
    
    # Conectar a BD
    print("\n🔌 Conectando a la base de datos...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor()
        print("✅ Conexión exitosa a Render PostgreSQL")
    except Exception as e:
        print(f"❌ Error de conexión: {e}")
        sys.exit(1)
    
    # PASO CRÍTICO: Crear tabla si no existe
    if not crear_tabla_si_no_existe(cursor):
        conn.close()
        sys.exit(1)
    
    # Cargar Excel
    print(f"\n📂 Cargando archivo {EXCEL_PATH}...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb['CRONOGRAMA PESV']
        print(f"✅ Excel cargado - Hoja: 'CRONOGRAMA PESV'")
        print(f"   📊 Dimensiones: {ws.max_row} filas, {ws.max_column} columnas")
    except Exception as e:
        print(f"❌ Error al cargar Excel: {e}")
        conn.close()
        sys.exit(1)
    
    # Limpiar tabla existente
    print("\n🗑️  Limpiando datos anteriores del plan anual...")
    try:
        cursor.execute("DELETE FROM plan_anual_trabajo")
        conn.commit()
        print("✅ Tabla limpia y lista para importar")
    except Exception as e:
        print(f"❌ Error al limpiar tabla: {e}")
        conn.close()
        sys.exit(1)
    
    registros_insertados = 0
    registros_con_error = 0
    
    # Mapeo de columnas del Excel (basado en la estructura que mencionaste)
    COL_ACTIVIDAD = 2      # B - Actividad
    COL_EVIDENCIA = 3      # C - Evidencia
    COL_CICLO = 4          # D - Ciclo PHVA
    COL_ARTICULOS = 5      # E - Artículos del Decreto
    COL_NIVEL = 6          # F - Nivel PESV
    COL_RESPONSABLES = 7   # G - Responsables
    COL_RECURSOS = 8       # H - Recursos
    COL_INICIO_MESES = 9   # I - Inicio de meses (Enero semana 1)
    
    print("\n🔄 Importando actividades...")
    print("   (Esto puede tomar 1-2 minutos...)")
    
    # Procesar desde fila 13 (después de los encabezados)
    for fila_num in range(13, ws.max_row + 1):
        row = ws[fila_num]
        
        # Extraer actividad (columna B)
        actividad = row[COL_ACTIVIDAD - 1].value
        if not actividad or str(actividad).strip() == '':
            continue  # Saltar filas vacías
        
        # Mostrar progreso
        if registros_insertados % 10 == 0 and registros_insertados > 0:
            print(f"   📊 Procesados: {registros_insertados} registros...")
        
        # Extraer datos básicos
        evidencia = row[COL_EVIDENCIA - 1].value
        ciclo_phva = row[COL_CICLO - 1].value
        articulos = row[COL_ARTICULOS - 1].value
        nivel = row[COL_NIVEL - 1].value
        responsables = row[COL_RESPONSABLES - 1].value
        recursos = row[COL_RECURSOS - 1].value
        
        # Extraer meses (12 meses × 4 semanas × 2 [P, E] = 96 columnas)
        meses_data = []
        col_actual = COL_INICIO_MESES - 1
        
        for mes in range(12):  # 12 meses
            for semana in range(4):  # 4 semanas por mes
                # P (Planificado)
                p_val = row[col_actual].value if col_actual < len(row) else None
                p = False
                if p_val is not None:
                    p_str = str(p_val).upper().strip()
                    p = p_str in ['P', 'E', 'X', 'SI', 'SÍ', '1', 'TRUE', 'V', 'VERDADERO']
                meses_data.append(p)
                col_actual += 1
                
                # E (Ejecutado)
                e_val = row[col_actual].value if col_actual < len(row) else None
                e = False
                if e_val is not None:
                    e_str = str(e_val).upper().strip()
                    e = e_str in ['E', 'X', 'SI', 'SÍ', '1', 'TRUE', 'V', 'VERDADERO']
                meses_data.append(e)
                col_actual += 1
        
        # Calcular estado y porcentaje
        semanas_planificadas = sum(1 for i in range(0, len(meses_data), 2) if meses_data[i])
        semanas_ejecutadas = sum(1 for i in range(1, len(meses_data), 2) if meses_data[i])
        
        # Determinar estado
        if semanas_planificadas == 0:
            estado = 'sin_planificar'
            porcentaje = 0.0
        elif semanas_ejecutadas >= semanas_planificadas:
            estado = 'completado'
            porcentaje = 100.0
        elif semanas_ejecutadas > 0:
            estado = 'en_proceso'
            porcentaje = round((semanas_ejecutadas / semanas_planificadas * 100), 2)
        else:
            estado = 'pendiente'
            porcentaje = 0.0
        
        # Insertar en la base de datos
        try:
            # Preparar los 96 valores booleanos para los meses
            params = [
                str(actividad)[:500] if actividad else '',
                str(evidencia)[:500] if evidencia else None,
                str(ciclo_phva)[:20] if ciclo_phva else None,
                str(articulos)[:200] if articulos else None,
                str(nivel)[:100] if nivel else None,
                str(responsables)[:200] if responsables else None,
                str(recursos)[:200] if recursos else None,
                *meses_data,  # Los 96 valores booleanos
                estado,
                porcentaje
            ]
            
            # Query de inserción (96 valores + 7 campos básicos + 2 campos de control = 105 parámetros)
            cursor.execute("""
                INSERT INTO plan_anual_trabajo (
                    actividad, evidencia, ciclo_phva, articulos_decreto, nivel_pesv,
                    responsables, recursos,
                    enero_semana1_p, enero_semana1_e, enero_semana2_p, enero_semana2_e,
                    enero_semana3_p, enero_semana3_e, enero_semana4_p, enero_semana4_e,
                    febrero_semana1_p, febrero_semana1_e, febrero_semana2_p, febrero_semana2_e,
                    febrero_semana3_p, febrero_semana3_e, febrero_semana4_p, febrero_semana4_e,
                    marzo_semana1_p, marzo_semana1_e, marzo_semana2_p, marzo_semana2_e,
                    marzo_semana3_p, marzo_semana3_e, marzo_semana4_p, marzo_semana4_e,
                    abril_semana1_p, abril_semana1_e, abril_semana2_p, abril_semana2_e,
                    abril_semana3_p, abril_semana3_e, abril_semana4_p, abril_semana4_e,
                    mayo_semana1_p, mayo_semana1_e, mayo_semana2_p, mayo_semana2_e,
                    mayo_semana3_p, mayo_semana3_e, mayo_semana4_p, mayo_semana4_e,
                    junio_semana1_p, junio_semana1_e, junio_semana2_p, junio_semana2_e,
                    junio_semana3_p, junio_semana3_e, junio_semana4_p, junio_semana4_e,
                    julio_semana1_p, julio_semana1_e, julio_semana2_p, julio_semana2_e,
                    julio_semana3_p, julio_semana3_e, julio_semana4_p, julio_semana4_e,
                    agosto_semana1_p, agosto_semana1_e, agosto_semana2_p, agosto_semana2_e,
                    agosto_semana3_p, agosto_semana3_e, agosto_semana4_p, agosto_semana4_e,
                    septiembre_semana1_p, septiembre_semana1_e, septiembre_semana2_p, septiembre_semana2_e,
                    septiembre_semana3_p, septiembre_semana3_e, septiembre_semana4_p, septiembre_semana4_e,
                    octubre_semana1_p, octubre_semana1_e, octubre_semana2_p, octubre_semana2_e,
                    octubre_semana3_p, octubre_semana3_e, octubre_semana4_p, octubre_semana4_e,
                    noviembre_semana1_p, noviembre_semana1_e, noviembre_semana2_p, noviembre_semana2_e,
                    noviembre_semana3_p, noviembre_semana3_e, noviembre_semana4_p, noviembre_semana4_e,
                    diciembre_semana1_p, diciembre_semana1_e, diciembre_semana2_p, diciembre_semana2_e,
                    diciembre_semana3_p, diciembre_semana3_e, diciembre_semana4_p, diciembre_semana4_e,
                    estado, porcentaje_avance
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s
                )
            """, params)
            
            registros_insertados += 1
            
            # Commit cada 20 registros para mejor rendimiento
            if registros_insertados % 20 == 0:
                conn.commit()
                
        except Exception as e:
            registros_con_error += 1
            error_msg = str(e)
            if "plan_anual_trabajo" in error_msg and "does not exist" in error_msg:
                print(f"\n❌ ERROR CRÍTICO: La tabla no existe aún después de intentar crearla")
                print(f"   Fila {fila_num}: {str(actividad)[:50]}...")
                break
            else:
                print(f"   ⚠️  Error en fila {fila_num}: {str(e)[:80]}...")
                if "column" in error_msg and "does not exist" in error_msg:
                    print(f"   💡 Posible problema de estructura de tabla")
            continue
    
    # Commit final
    conn.commit()
    
    # Verificar cuántos registros quedaron
    cursor.execute("SELECT COUNT(*) FROM plan_anual_trabajo")
    total_en_bd = cursor.fetchone()[0]
    
    # Mostrar resumen detallado
    if total_en_bd > 0:
        cursor.execute("""
            SELECT estado, COUNT(*) as cantidad, 
                   ROUND(AVG(porcentaje_avance), 2) as promedio_avance
            FROM plan_anual_trabajo
            GROUP BY estado
            ORDER BY estado
        """)
        estadisticas = cursor.fetchall()
        
        cursor.execute("SELECT DISTINCT ciclo_phva FROM plan_anual_trabajo WHERE ciclo_phva IS NOT NULL")
        ciclos = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    print("\n" + "=" * 60)
    print("✅ IMPORTACIÓN COMPLETADA - RESUMEN")
    print("=" * 60)
    print(f"   ✔️  Registros insertados: {registros_insertados}")
    print(f"   ⚠️  Registros con error: {registros_con_error}")
    print(f"   📊 Total en base de datos: {total_en_bd}")
    
    if total_en_bd > 0 and estadisticas:
        print(f"\n📈 Estadísticas por estado:")
        for estado, cantidad, promedio in estadisticas:
            print(f"   • {estado}: {cantidad} actividades ({promedio}% avance)")
        
        if ciclos:
            print(f"\n🎯 Ciclos PHVA encontrados: {', '.join([c[0] for c in ciclos if c[0]])}")
    
    print("\n🎉 ¡Listo! Ahora puedes ver las actividades en:")
    print("   👉 http://tu-servidor/sst/plan-anual")
    print("=" * 60)

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Importación cancelada por el usuario")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Error fatal: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
